# ══════════════════════════════════════════════════════════════════
#  ★ NEW: 라운드별 입력 데이터 시각화 Excel
#   · 첨부 이미지 형식 — ① POD 분포(grid)  ② Weight 분포(grid)  ③ 선적순서 리스트
#   · ★ 입력 컨테이너는 정책과 무관 (env.reset 의 seed 만 의존) → 1회만 생성
#   · 표시 대상: 각 라운드의 "대표 1개 에피소드" (eval ep=0, eval_seed=GLOBAL_SEED)
#   · grid 배치는 env 의 적재 결과가 아닌 "입력 순서를 그대로 채운 참조 배치"가 아니라,
#     실제 SF 정책 최종 bay plan(per_round_bay_plans)을 사용해 적재 위치를 표현
# ══════════════════════════════════════════════════════════════════
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── POD 색상 팔레트 (env / 기존 Excel 과 동일) ──
POD_FILLS = {
    1: "4393C3", 2: "F4A582", 3: "D6604D",
    4: "92C5DE", 5: "2166AC", 6: "B2182B",
}
POD_FONT_WHITE = {3, 5, 6}   # 진한 배경 → 흰 글씨

# ── Weight heatmap 색상 (저=초록 / 중=노랑 / 고=빨강) ──
def _wt_fill(w, wmin, wmax):
    if wmax - wmin < 1e-6:
        return "C6EFCE"
    t = (w - wmin) / (wmax - wmin)
    if t < 0.34:   return "C6EFCE"   # green
    elif t < 0.67: return "FFEB9C"   # yellow
    else:          return "FFC7CE"   # red

hdr_font   = Font(bold=True, color="FFFFFF", size=10)
blue_fill  = PatternFill("solid", fgColor="4472C4")
sect_font  = Font(bold=True, size=11)
center     = Alignment(horizontal="center", vertical="center")
thin       = Border(*[Side(style="thin")] * 4)
green_font = Font(bold=True, color="008000", size=10)

def _bay_plan_from_env(seed_for_ep, level, config):
    """대표 에피소드의 입력 컨테이너 + (정책 적재 없이) 입력 순서 메타 재생성.

    실제 적재 grid 는 SF 정책 결과(per_round_bay_plans)를 사용하고,
    선적순서 리스트는 env.reset 의 deterministic 시퀀스로 재생성한다.
    """
    env = make_env(level=level, config=config)
    env.reset(seed=seed_for_ep)
    return env.ctns_pod.copy(), env.ctns_wt.copy(), env.n_rows, env.n_tiers

wb = Workbook()
wb.remove(wb.active)

# 입력 데이터는 정책 무관 → SF 결과의 bay plan(적재 위치)을 표현용으로 사용
ref_res = all_results.get("SF", next(iter(all_results.values())))

for rnd_idx in range(CONFIG["total_rounds"]):
    lv     = ref_res["levels"][rnd_idx]
    lv_cfg = CONFIG["levels"][lv]
    n_rows, n_tiers = lv_cfg["n_rows"], lv_cfg["n_tiers"]

    # 대표 에피소드 = eval ep0 (seed = GLOBAL_SEED + 0*17 = GLOBAL_SEED)
    rep_seed = GLOBAL_SEED + 0 * 17
    ctns_pod, ctns_wt, _, _ = _bay_plan_from_env(rep_seed, lv, CONFIG)

    # 실제 적재 grid (SF 정책 최종 라운드 bay plan)
    bp = ref_res["per_round_bay_plans"][rnd_idx]
    pod_grid = bp["pod_grid"]     # (n_rows, n_tiers)
    wt_grid  = bp["wt_grid"]
    stack_h  = bp["stack_h"]

    ws = wb.create_sheet(f"R{rnd_idx+1}_Lv{lv}")
    ws.sheet_view.showGridLines = False

    # ── 제목 ──
    ncol = n_rows + 2
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=ncol)
    t = ws.cell(row=1, column=2,
                value=f"Round {rnd_idx+1} (Lv{lv}) — Bay Plan: {n_rows}R × {n_tiers}T")
    t.font = Font(bold=True, size=13); t.alignment = center
    t.fill = PatternFill("solid", fgColor="FFFFFF")
    t.border = Border(top=Side(style="medium"), bottom=Side(style="medium"),
                      left=Side(style="medium"), right=Side(style="medium"))

    r0 = 3   # 현재 작성 행 포인터

    # ════════ ① POD 분포 ════════
    ws.cell(row=r0, column=1, value="【POD 분포】").font = sect_font
    r0 += 1
    # 헤더: Tier\Row | R0..Rn | Col ΣWT
    c = ws.cell(row=r0, column=1, value="Tier\\Row"); c.font = hdr_font; c.fill = blue_fill; c.alignment = center; c.border = thin
    for rr in range(n_rows):
        c = ws.cell(row=r0, column=2 + rr, value=f"R{rr}")
        c.font = hdr_font; c.fill = blue_fill; c.alignment = center; c.border = thin
    c = ws.cell(row=r0, column=2 + n_rows, value="Col ΣWT")
    c.font = hdr_font; c.fill = PatternFill("solid", fgColor="8FAADC"); c.alignment = center; c.border = thin
    r0 += 1
    # Tier 행 (위→아래: 최상단 tier 부터)
    pod_top_row = r0
    for t_disp in range(n_tiers - 1, -1, -1):
        c = ws.cell(row=r0, column=1, value=f"T{t_disp}")
        c.font = hdr_font; c.fill = blue_fill; c.alignment = center; c.border = thin
        for rr in range(n_rows):
            pod = int(pod_grid[rr, t_disp]) if t_disp < int(stack_h[rr]) else 0
            cell = ws.cell(row=r0, column=2 + rr)
            cell.alignment = center; cell.border = thin
            if pod > 0:
                cell.value = f"{POD_NAMES[pod][:3]}({pod})"
                cell.fill  = PatternFill("solid", fgColor=POD_FILLS[pod])
                cell.font  = Font(bold=True, size=9,
                                  color="FFFFFF" if pod in POD_FONT_WHITE else "000000")
            else:
                cell.value = "—"
                cell.fill  = PatternFill("solid", fgColor="EFEFEF")
                cell.font  = Font(size=9, color="808080")
        r0 += 1
    # Col ΣWT 행
    c = ws.cell(row=r0, column=1, value="Col ΣWT")
    c.font = hdr_font; c.fill = blue_fill; c.alignment = center; c.border = thin
    for rr in range(n_rows):
        col_wt = float(wt_grid[rr, :int(stack_h[rr])].sum())
        cc = ws.cell(row=r0, column=2 + rr, value=round(col_wt, 1))
        cc.font = green_font; cc.alignment = center; cc.border = thin
    r0 += 2

    # ════════ ② Weight 분포 (MT) ════════
    ws.cell(row=r0, column=1, value="【Weight 분포 (MT)】").font = sect_font
    r0 += 1
    c = ws.cell(row=r0, column=1, value="Tier\\Row"); c.font = hdr_font; c.fill = blue_fill; c.alignment = center; c.border = thin
    for rr in range(n_rows):
        c = ws.cell(row=r0, column=2 + rr, value=f"R{rr}")
        c.font = hdr_font; c.fill = blue_fill; c.alignment = center; c.border = thin
    r0 += 1
    # heatmap 범위
    occ_wts = [float(wt_grid[rr, t]) for rr in range(n_rows)
               for t in range(int(stack_h[rr]))]
    wmin = min(occ_wts) if occ_wts else 0.0
    wmax = max(occ_wts) if occ_wts else 1.0
    for t_disp in range(n_tiers - 1, -1, -1):
        c = ws.cell(row=r0, column=1, value=f"T{t_disp}")
        c.font = hdr_font; c.fill = blue_fill; c.alignment = center; c.border = thin
        for rr in range(n_rows):
            cell = ws.cell(row=r0, column=2 + rr)
            cell.alignment = center; cell.border = thin
            if t_disp < int(stack_h[rr]):
                w = float(wt_grid[rr, t_disp])
                cell.value = round(w, 1)
                cell.fill  = PatternFill("solid", fgColor=_wt_fill(w, wmin, wmax))
                cell.font  = Font(size=9)
            else:
                cell.value = "—"
                cell.fill  = PatternFill("solid", fgColor="EFEFEF")
                cell.font  = Font(size=9, color="808080")
        r0 += 1
    r0 += 2

    # ════════ ③ 선적 순서 리스트 ════════
    ws.cell(row=r0, column=1, value="【선적 순서 리스트】").font = sect_font
    r0 += 1
    list_hdrs = ["순번", "POD ID", "POD Name", "Weight (MT)"]
    for col, h in enumerate(list_hdrs, 1):
        c = ws.cell(row=r0, column=col, value=h)
        c.font = hdr_font; c.fill = blue_fill; c.alignment = center; c.border = thin
    r0 += 1
    for i in range(len(ctns_pod)):
        pod = int(ctns_pod[i]); w = float(ctns_wt[i])
        ws.cell(row=r0, column=1, value=i + 1).alignment = center
        ws.cell(row=r0, column=2, value=pod).alignment = center
        ws.cell(row=r0, column=3, value=POD_NAMES[pod]).alignment = center
        ws.cell(row=r0, column=4, value=round(w, 1)).alignment = center
        for col in range(1, 5):
            ws.cell(row=r0, column=col).border = thin
        r0 += 1

    # ── 컬럼 너비 ──
    ws.column_dimensions["A"].width = 12
    for rr in range(n_rows):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(2 + rr)].width = 11
    from openpyxl.utils import get_column_letter
    ws.column_dimensions[get_column_letter(2 + n_rows)].width = 10

dist_path = f"/content/results/{CONFIG['experiment_name']}_BayPlan_Distributions_seed{GLOBAL_SEED}.xlsx"
wb.save(dist_path)
print("═" * 70)
print("  📊 라운드별 분포 Excel 저장 완료")
print("═" * 70)
print(f"  파일: {dist_path}")
_sheet_names = [f"R{i+1}_Lv" + str(ref_res["levels"][i]) for i in range(CONFIG["total_rounds"])]
print(f"  시트: {_sheet_names}")
print(f"  내용: ① POD 분포 grid  ② Weight 분포 heatmap  ③ 선적순서 리스트")
print(f"  ※ 적재 grid = SF 정책 최종 bay plan / 선적순서 = deterministic 입력 시퀀스")

try:
    from google.colab import files
    files.download(dist_path)
    print("  ⬇️  Colab 다운로드 트리거 완료")
except Exception:
    print(f"  ℹ️  (비Colab) 저장 경로에서 직접 확인: {dist_path}")
