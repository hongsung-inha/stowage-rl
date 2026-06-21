# v17 6-policy 비교 Excel 저장 (대표 3정책 BAL/STAB/EFF KPI 포함)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def save_v13_comparison_excel(all_results, all_loggers, base_log_dir):
    wb = Workbook()
    HF = Font(bold=True, color="FFFFFF", size=10)
    HFILL = PatternFill("solid", fgColor="1F3864")
    CTR = Alignment(horizontal="center", vertical="center")

    # Sheet 1: Round_Summary (6 policies × 4 라운드)
    ws1 = wb.active; ws1.title = "Round_Summary"
    headers = ["Policy", "Round", "Level", "Phase", "Mode",
               "Reward", "Std", "OSR", "VPR", "WBI", "PSR", "CWVR",
               "Episodes", "Steps", "TrainTime(min)"]
    for c, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=c, value=h)
        cell.font, cell.fill, cell.alignment = HF, HFILL, CTR

    r = 2
    for pn, res in all_results.items():
        for rnd in range(CONFIG["total_rounds"]):
            pm = res["per_round_metrics"][rnd]
            pi = res["phase_info"][rnd]
            row = [
                pn, rnd+1, f"Lv{pm['level']}", pi["phase"], pi["mode"],
                round(pm["avg_reward"], 2), round(pm["std_reward"], 2),
                round(pm["avg_osr"], 4), round(pm["avg_vpr"], 4),
                round(pm["avg_wbi"], 4), round(pm["avg_psr"], 4),
                round(pm.get("avg_cwvr", 0), 4),
                pm.get("actual_episodes", 0), pm.get("actual_steps", 0),
                round(pm.get("train_seconds", 0)/60, 1),
            ]
            for c, v in enumerate(row, 1):
                ws1.cell(row=r, column=c, value=v).alignment = CTR
            r += 1

    # Sheet 2: Lv4 — 대표 3정책 (BAL/STAB/EFF) 비교
    ws2 = wb.create_sheet("Lv4_ThreeWay")
    ws2.cell(row=1, column=1, value="Final Round (Lv4) — 대표 3정책 BAL/STAB/EFF").font = Font(bold=True, size=12)
    ws2.merge_cells("A1:F1")
    for c, h in enumerate(["Metric", "BAL", "STAB", "EFF", "Best", "Direction"], 1):
        cell = ws2.cell(row=3, column=c, value=h)
        cell.font, cell.fill, cell.alignment = HF, HFILL, CTR

    KPIS = [
        ("Reward",       "avg_reward", "high"),
        ("VPR",          "avg_vpr",    "high"),
        ("WBI",          "avg_wbi",    "high"),
        ("OSR",          "avg_osr",    "low"),
        ("PSR",          "avg_psr",    "low"),
        ("CWVR",         "avg_cwvr",   "low"),
    ]
    for i, (name, key, direction) in enumerate(KPIS):
        bl_v = all_results[BAL_POLICY]["per_round_metrics"][-1].get(key, 0)
        sf_v = all_results[STAB_POLICY]["per_round_metrics"][-1].get(key, 0)
        ef_v = all_results[EFF_POLICY]["per_round_metrics"][-1].get(key, 0)
        vals = {"BAL": bl_v, "STAB": sf_v, "EFF": ef_v}
        if direction == "high":
            best = max(vals, key=vals.get)
        else:
            best = min(vals, key=vals.get)
        ws2.cell(row=4+i, column=1, value=name).font = Font(bold=True)
        ws2.cell(row=4+i, column=2, value=round(bl_v, 4)).alignment = CTR
        ws2.cell(row=4+i, column=3, value=round(sf_v, 4)).alignment = CTR
        ws2.cell(row=4+i, column=4, value=round(ef_v, 4)).alignment = CTR
        wc = ws2.cell(row=4+i, column=5, value=best)
        wc.font = Font(bold=True, color="C00000"); wc.alignment = CTR
        ws2.cell(row=4+i, column=6, value="↑ better" if direction=="high" else "↓ better").alignment = CTR

    # Sheet 3: H5 Tests (★ v16: vs Baseline 가설 검증)
    ws_h5 = wb.create_sheet("H5_Tests")
    ws_h5.cell(row=1, column=1, value="H5: 우선순위 부여 효과 (vs Baseline)").font = Font(bold=True, size=12)
    ws_h5.merge_cells("A1:F1")
    for c, h in enumerate(["Hypothesis", "KPI", "Policy", "Policy Val", "BL Val", "Δ (vs BL)", "Threshold", "Pass"], 1):
        cell = ws_h5.cell(row=3, column=c, value=h)
        cell.font, cell.fill, cell.alignment = HF, HFILL, CTR

    if True:
        bl_lv4 = all_results[BAL_POLICY]["per_round_metrics"][-1]
        sf_lv4 = all_results[STAB_POLICY]["per_round_metrics"][-1]
        ef_lv4 = all_results[EFF_POLICY]["per_round_metrics"][-1]
        h5_tests = [
            ("H5a", "WBI", "STAB", sf_lv4["avg_wbi"], bl_lv4["avg_wbi"],
             sf_lv4["avg_wbi"] - bl_lv4["avg_wbi"], "+0.02"),
            ("H5b", "OSR", "EFF", ef_lv4["avg_osr"], bl_lv4["avg_osr"],
             bl_lv4["avg_osr"] - ef_lv4["avg_osr"], "+0.02"),
        ]
        for i, (hyp, kpi, pol, pv, blv, delta, thr) in enumerate(h5_tests):
            passed = "✓ Pass" if delta > 0.02 else "✗ Fail"
            row = [hyp, kpi, pol, round(pv, 4), round(blv, 4), round(delta, 4), thr, passed]
            for c, v in enumerate(row, 1):
                cell = ws_h5.cell(row=4+i, column=c, value=v)
                cell.alignment = CTR
                if c == 8:
                    cell.font = Font(bold=True, color="00B050" if "Pass" in passed else "C00000")

    # Sheet 4: Phase_Info (정책별 phase 별 활성 보상)
    ws3 = wb.create_sheet("Phase_Info")
    for c, h in enumerate(["Policy", "Round", "Level", "Phase", "Mode",
                            "N_Active_Rewards"], 1):
        cell = ws3.cell(row=1, column=c, value=h)
        cell.font, cell.fill, cell.alignment = HF, HFILL, CTR
    r = 2
    for pn, res in all_results.items():
        for pi in res["phase_info"]:
            row = [pn, pi["round"], f"Lv{pi['level']}", pi["phase"], pi["mode"],
                   pi["n_active_rewards"]]
            for c, v in enumerate(row, 1):
                ws3.cell(row=r, column=c, value=v).alignment = CTR
            r += 1

    # Sheet 5: Config
    ws4 = wb.create_sheet("Config")
    items = [
        ("Version", "v17 6-policy 안정:효율 비중 가중 (ES/SF_9_EF_1/SF_7_EF_3/BL_SF_5_EF_5/EF_7_SF_3/EF_9_SF_1)"),
        ("Seed", GLOBAL_SEED),
        ("N Policies", len(POLICIES)),
        ("Total Episodes/Policy", sum(CONFIG["episodes_per_level"].values())),
        ("Total Episodes (All)", sum(CONFIG["episodes_per_level"].values()) * len(POLICIES)),
        ("Rounds (all policies)", str(POLICIES[REF_POLICY]["phase1_rounds"] + POLICIES[REF_POLICY]["phase2_rounds"])),
        ("STAB Policy / Mode", f"{STAB_POLICY} / {POLICIES[STAB_POLICY]['phase1_mode']}"),
        ("EFF Policy / Mode",  f"{EFF_POLICY} / {POLICIES[EFF_POLICY]['phase1_mode']}"),
        ("BAL Policy / Mode",  f"{BAL_POLICY} / {POLICIES[BAL_POLICY]['phase1_mode']}"),
        ("All Policies", ", ".join(POLICIES.keys())),
        ("PPO LR", CONFIG["ppo"]["lr"]),
        ("PPO net_arch", str(CONFIG["ppo"]["net_arch"])),
        ("R5 weight_balance",     CONFIG["rw"]["weight_balance"]),
        ("R6 cog_penalty",        CONFIG["rw"]["cog_penalty"]),
        ("R11 wt_inversion_penalty", CONFIG["rw"]["wt_inversion_penalty"]),
    ]
    for i, (k, v) in enumerate(items, 1):
        ws4.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws4.cell(row=i, column=2, value=str(v))
    ws4.column_dimensions["A"].width = 28
    ws4.column_dimensions["B"].width = 56

    out = f"{base_log_dir}/v17_6pol_results.xlsx"
    wb.save(out)
    return out


# ── Excel + ZIP 다운로드 ──
print("📊 Saving v17 6-policy comparison Excel...")
xlsx_path = save_v13_comparison_excel(all_results, all_loggers,
                                       all_loggers[REF_POLICY].log_dir)
print(f"  💾 {xlsx_path}")

# 정책별 개별 Excel (v8 형식)
print()
print("📊 Saving per-policy Excels (v16 format)...")
per_pol_xlsx = []
for pn in POLICIES:
    cfg = copy.deepcopy(CONFIG)
    cfg["experiment_name"] = f"v17_6pol_{pn}"
    p = save_results_to_excel(all_results[pn], cfg, GLOBAL_SEED, all_loggers[pn])
    per_pol_xlsx.append(p)

# ZIP 묶기
import shutil
zip_dir = "/content/v17_6pol_results"
os.makedirs(zip_dir, exist_ok=True)
for pn, lg in all_loggers.items():
    sub = os.path.join(zip_dir, pn)
    os.makedirs(sub, exist_ok=True)
    for f in os.listdir(lg.log_dir):
        src = os.path.join(lg.log_dir, f)
        if os.path.isfile(src):
            shutil.copy2(src, sub)

zip_path = "/content/v17_6pol_results.zip"
shutil.make_archive(zip_path.replace(".zip", ""), "zip", zip_dir)
print(f"\n📦 Bundle: {zip_path}")

# 다운로드
print()
print("⬇  Downloading...")
from google.colab import files
files.download(xlsx_path)
files.download(zip_path)

print()
print("=" * 70)
print("  ✅ v17 6-policy Pilot 완료")
print(f"  최종 누적 보상 (Step 기준):")
for pn, fr in final_cum_rewards.items():
    print(f"    {pn}: {fr:+,.0f}")
print(f"  가설 검증 결과는 Cell 9 출력 참조")
print("=" * 70)
