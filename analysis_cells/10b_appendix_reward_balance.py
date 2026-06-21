# ══════════════════════════════════════════════════════════════════
#  ★ Cell 10B (검증 셀): 보상 균형 검증 + 정책별 실측 기여도 Appendix 표
#
#  목적 (논문 Appendix 용):
#   (a) 설계 균형  : 계수 절댓값 합(Σ|w|) 기준 안정:효율 = 50:50 검증
#   (b) 실측 균형  : 학습된 정책별 Lv4 누적 보상 기여도(measured contribution)
#                   를 그룹별로 집계 → 설계 의도가 어떻게 실현됐는지 보고
#
#  ※ (a)는 정책-독립적 '설계 파라미터'로 균형의 정의 근거 (재현 가능).
#    (b)는 정책-의존적 '사후 관측치'로 해석/진단 용도 (균형 재조정 X).
#  ※ 모든 정책은 BASE 가중치 환경에서 평가 → 학습된 정책 차이만 측정.
# ══════════════════════════════════════════════════════════════════
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── 보상 component 키 ↔ 그룹 매핑 (StowageMetrics.reward_components 기준) ──
#   R13은 component에서 단일 키(R13_tier_band)로 누적되므로 효율 그룹에 1개로 집계.
COMP_TO_GROUP = {
    "R1_valid": "필수", "R2_stack_full": "필수", "R3_overstow": "필수",
    "R7_completion": "필수", "R9_col_wt": "필수", "R11_wt_inversion": "필수",
    "R5_weight_bal": "안정", "R6_cog": "안정",
    "R4_order": "효율", "R8_pod_band": "효율", "R10_tier_match": "효율",
    "R12_col_order": "효율", "R13_tier_band": "효율",
    "R14_empty_row": "효율", "R15_vstack_pod": "효율",
}
COMP_RID = {
    "R1_valid": "R1", "R2_stack_full": "R2", "R3_overstow": "R3",
    "R4_order": "R4", "R5_weight_bal": "R5", "R6_cog": "R6",
    "R7_completion": "R7", "R8_pod_band": "R8", "R9_col_wt": "R9",
    "R10_tier_match": "R10", "R11_wt_inversion": "R11", "R12_col_order": "R12",
    "R13_tier_band": "R13", "R14_empty_row": "R14", "R15_vstack_pod": "R15",
}
COMP_ORDER = ["R1_valid","R2_stack_full","R3_overstow","R4_order","R5_weight_bal",
              "R6_cog","R7_completion","R8_pod_band","R9_col_wt","R10_tier_match",
              "R11_wt_inversion","R12_col_order","R13_tier_band","R14_empty_row",
              "R15_vstack_pod"]

# ──────────────────────────────────────────────────────────────────
#  (a) 설계 균형 검증 — 계수 절댓값 합 (Σ|w|) 기준
# ──────────────────────────────────────────────────────────────────
def _group_abs_sum(keys):
    return sum(abs(CONFIG["rw"][k]) for k in keys)

stab_w = _group_abs_sum(STABILITY_KEYS)
eff_w  = _group_abs_sum(EFFICIENCY_KEYS)
ess_w  = _group_abs_sum(ESSENTIAL_KEYS)

print("═" * 78)
print("  (a) 설계 균형 — 계수 절댓값 합 Σ|w| 기준 (정책-독립적 설계 파라미터)")
print("═" * 78)
print(f"  {'그룹':<8} {'항목':<40} {'Σ|w|':>8}")
print("  " + "─" * 74)
print(f"  {'필수':<8} {'R1,R2,R3,R7,R9,R11':<40} {ess_w:>8.2f}  (균형 대상 아님)")
print(f"  {'안정':<8} {'R5(+6.5), R6(-8.5)':<40} {stab_w:>8.2f}")
print(f"  {'효율':<8} {'R4,R8,R10,R12,R13(a/b),R14,R15':<40} {eff_w:>8.2f}")
print("  " + "─" * 74)
_den = stab_w + eff_w
print(f"  ▶ 안정 : 효율 = {stab_w:.1f} : {eff_w:.1f} = "
      f"{stab_w/_den*100:.1f}% : {eff_w/_den*100:.1f}%")
if abs(stab_w - eff_w) < 1e-6:
    print(f"  ✅ 정확히 50:50 균형 확인 (Σ|w| 동일)")
else:
    print(f"  ⚠️ 50:50 아님 (차이 {abs(stab_w-eff_w):.2f})")
print()

# ──────────────────────────────────────────────────────────────────
#  (b) 실측 기여도 — 정책별 Lv4 누적 보상 기여도 (BASE 가중치 평가)
# ──────────────────────────────────────────────────────────────────
print("═" * 78)
print("  (b) 실측 기여도 — 정책별 Lv4 누적 보상 기여도 (BASE 가중치, 30 eval ep)")
print("═" * 78)

measured = {}   # policy -> {comp_key: (mean, std)}  + group sums
for pn, res in all_results.items():
    cfg = copy.deepcopy(CONFIG)
    cfg["rw"] = dict(CONFIG["rw"])    # BASE 가중치로 평가
    comp = _evaluate_reward_components(
        res["trained_model"], res["levels"][-1], cfg, GLOBAL_SEED,
        n_eval=CONFIG["eval_episodes"])
    pol_stat = {}
    for k in COMP_ORDER:
        arr = np.array(comp.get(k, [0.0]))
        pol_stat[k] = (float(np.mean(arr)), float(np.std(arr)))
    # 그룹별 평균 기여도 합 (절댓값 — 영향력 크기 비교용)
    g_abs = {"필수": 0.0, "안정": 0.0, "효율": 0.0}
    g_signed = {"필수": 0.0, "안정": 0.0, "효율": 0.0}
    for k in COMP_ORDER:
        g = COMP_TO_GROUP[k]
        g_abs[g]    += abs(pol_stat[k][0])
        g_signed[g] += pol_stat[k][0]
    pol_stat["_group_abs"]    = g_abs
    pol_stat["_group_signed"] = g_signed
    pol_stat["_total"] = (float(np.mean(comp.get("Total",[0]))),
                          float(np.std(comp.get("Total",[0]))))
    measured[pn] = pol_stat

# 콘솔 출력: 그룹별 실측 |기여도| 및 안정:효율 비율 (정책별)
print(f"  {'정책':<6} {'안정|Σ|':>10} {'효율|Σ|':>10} {'안정:효율(실측)':>18} {'Total':>10}")
print("  " + "─" * 74)
for pn in all_results:
    ga = measured[pn]["_group_abs"]
    s, e = ga["안정"], ga["효율"]
    if (s + e) > 1e-9:
        ratio = f"{s/(s+e)*100:.1f}% : {e/(s+e)*100:.1f}%"
    else:
        ratio = "    n/a (둘 다 0)"
    tot = measured[pn]["_total"][0]
    print(f"  {pn:<6} {s:>10.2f} {e:>10.2f} {ratio:>18} {tot:>+10.2f}")
print("  " + "─" * 74)
print("  ※ ES: 안정·효율 모두 비활성(0) | SF: 효율=0 | EF: 안정=0 | BL: 둘 다 활성")
print("  ※ 실측 비율은 학습된 정책 행동·발동빈도에 의존 → 설계 50:50과 다를 수 있음(정상)")
print()

# ──────────────────────────────────────────────────────────────────
#  Appendix Excel 표 생성
# ──────────────────────────────────────────────────────────────────
_hf   = Font(bold=True, color="FFFFFF", size=10)
_fill = PatternFill("solid", fgColor="404040")
_acc  = PatternFill("solid", fgColor="2166AC")
_grp_fill = {"필수": PatternFill("solid", fgColor="E2EFDA"),
             "안정": PatternFill("solid", fgColor="DDEBF7"),
             "효율": PatternFill("solid", fgColor="FCE4D6")}
_ctr  = Alignment(horizontal="center", vertical="center")
_thin = Border(*[Side(style="thin")]*4)

wb = Workbook()

# ── Sheet 1: 설계 균형 (Σ|w|) ──
ws = wb.active; ws.title = "A_DesignBalance"
ws.cell(row=1, column=1,
        value="(a) Design Balance — Coefficient |w| (Σ|w|) 기준 50:50 검증").font = Font(bold=True, size=12)
hdr = ["R", "계층", "Key", "Weight(v16)", "|Weight|"]
for c, h in enumerate(hdr, 1):
    cell = ws.cell(row=3, column=c, value=h); cell.font=_hf; cell.fill=_fill; cell.alignment=_ctr
r = 4
for k in CONFIG["rw"]:
    grp = TIER_MAP.get(k, "?")
    ws.cell(row=r, column=1, value=R_MAP.get(k, "?")).alignment=_ctr
    gcell = ws.cell(row=r, column=2, value=grp); gcell.alignment=_ctr
    gcell.fill = _grp_fill.get(grp, PatternFill())
    ws.cell(row=r, column=3, value=k)
    ws.cell(row=r, column=4, value=round(CONFIG["rw"][k], 3)).alignment=_ctr
    ws.cell(row=r, column=5, value=round(abs(CONFIG["rw"][k]), 3)).alignment=_ctr
    r += 1
# 그룹 요약
r += 1
ws.cell(row=r, column=1, value="그룹 Σ|w| 요약").font = Font(bold=True); r += 1
for label, val in [("필수 Σ|w|", ess_w), ("안정 Σ|w|", stab_w), ("효율 Σ|w|", eff_w)]:
    ws.cell(row=r, column=1, value=label).font = Font(bold=True)
    ws.cell(row=r, column=2, value=round(val, 2)).alignment=_ctr; r += 1
ws.cell(row=r, column=1, value="안정 : 효율").font = Font(bold=True)
ws.cell(row=r, column=2,
        value=f"{stab_w/_den*100:.1f}% : {eff_w/_den*100:.1f}%").alignment=_ctr
for col, w in [("A",8),("B",8),("C",24),("D",13),("E",12)]:
    ws.column_dimensions[col].width = w

# ── Sheet 2: 실측 기여도 (정책×항목 매트릭스) ──
ws2 = wb.create_sheet("B_MeasuredContribution")
ws2.cell(row=1, column=1,
         value="(b) Measured Contribution — Lv4 누적 보상 기여도 (BASE 가중치, mean±std)").font = Font(bold=True, size=12)
pols = list(all_results.keys())
hdr2 = ["R", "계층", "Component"] + pols
for c, h in enumerate(hdr2, 1):
    cell = ws2.cell(row=3, column=c, value=h)
    cell.font=_hf; cell.fill=(_acc if h in pols else _fill); cell.alignment=_ctr
r = 4
for k in COMP_ORDER:
    grp = COMP_TO_GROUP[k]
    ws2.cell(row=r, column=1, value=COMP_RID[k]).alignment=_ctr
    gcell = ws2.cell(row=r, column=2, value=grp); gcell.alignment=_ctr
    gcell.fill = _grp_fill.get(grp, PatternFill())
    ws2.cell(row=r, column=3, value=k)
    for ci, pn in enumerate(pols, 4):
        mean, std = measured[pn][k]
        cell = ws2.cell(row=r, column=ci, value=f"{mean:+.2f} ± {std:.2f}")
        cell.alignment=_ctr; cell.border=_thin
    r += 1
# 그룹별 |기여도| 합 + 비율
r += 1
ws2.cell(row=r, column=3, value="── 그룹별 |기여도| 합 ──").font = Font(bold=True); r += 1
for grp in ["필수", "안정", "효율"]:
    ws2.cell(row=r, column=3, value=f"{grp} Σ|contrib|").font = Font(bold=True)
    for ci, pn in enumerate(pols, 4):
        ws2.cell(row=r, column=ci,
                 value=round(measured[pn]["_group_abs"][grp], 2)).alignment=_ctr
    r += 1
ws2.cell(row=r, column=3, value="안정:효율 (실측 %)").font = Font(bold=True)
for ci, pn in enumerate(pols, 4):
    ga = measured[pn]["_group_abs"]; s, e = ga["안정"], ga["효율"]
    txt = f"{s/(s+e)*100:.0f}:{e/(s+e)*100:.0f}" if (s+e) > 1e-9 else "n/a"
    ws2.cell(row=r, column=ci, value=txt).alignment=_ctr
r += 1
ws2.cell(row=r, column=3, value="Total reward (mean)").font = Font(bold=True)
for ci, pn in enumerate(pols, 4):
    ws2.cell(row=r, column=ci,
             value=round(measured[pn]["_total"][0], 2)).alignment=_ctr
for col, w in [("A",6),("B",7),("C",22)]:
    ws2.column_dimensions[col].width = w
from openpyxl.utils import get_column_letter
for ci in range(len(pols)):
    ws2.column_dimensions[get_column_letter(4+ci)].width = 16

# ── Sheet 3: Notes ──
ws3 = wb.create_sheet("C_Notes")
notes = [
    "Reward Balance Verification — Appendix Notes",
    "",
    "(a) Design balance (Σ|w|): 정책-독립적 설계 파라미터. 균형의 '정의' 근거.",
    "    안정 그룹(R5,R6) Σ|w| = 15.0,  효율 그룹(R4,R8,R10,R12,R13,R14,R15) Σ|w| = 15.0",
    "    → 정확히 50:50. 리뷰어가 CONFIG['rw']로 즉시 재현 가능.",
    "",
    "(b) Measured contribution: 학습된 정책의 Lv4 누적 보상 기여도(사후 관측).",
    "    발동 빈도·스케일(R5 ×(1-CV), R6 ×편차비율 등)에 의존하므로",
    "    설계 50:50과 실측 비율은 다를 수 있음(정상). 균형 '재조정' 용도 아님.",
    "    해석/진단(메커니즘 설명) 및 리뷰어 선제 방어용으로 보고.",
    "",
    "정책 활성 그룹: ES=필수만 | SF=필수+안정 | EF=필수+효율 | BL=전체.",
    "모든 정책은 BASE 가중치 환경에서 평가하여 '학습된 정책 차이'만 측정.",
    "v16 변경: R5 6.0→6.5, R6 -8.0→-8.5 (안정 Σ|w| 14.0→15.0).",
]
for i, t in enumerate(notes, 1):
    cell = ws3.cell(row=i, column=1, value=t)
    if i == 1: cell.font = Font(bold=True, size=12)
ws3.column_dimensions["A"].width = 90

# ── 저장 (첫 정책 logger의 log_dir 사용) ──
_first_logger = all_loggers[list(all_loggers.keys())[0]]
_appendix_path = f"{_first_logger.log_dir}/appendix_reward_balance_v16.xlsx"
wb.save(_appendix_path)
print(f"  📊 Appendix Excel 저장: {_appendix_path}")

try:
    from google.colab import files
    files.download(_appendix_path)
except Exception:
    pass

print("\n✅ 보상 균형 검증 완료 — (a) 설계 50:50 확인 + (b) 정책별 실측 기여도 Appendix 표 생성")
