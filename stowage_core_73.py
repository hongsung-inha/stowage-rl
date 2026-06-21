# ══════════════════════════════════════════════════════════════════════
#  stowage_core_73.py — 7:3 비율 실험 공통 모듈
#  (TRAIN/AGGREGATE 노트북 Cell 5 정의 + Cell 7 CONFIG/POLICIES 추출)
#  · OUTPUT_BASE/RESULTS_DIR : VESSL_OUTPUT_DIR → /shared → ./results_vessl
#  · CBData : TRAIN(저장)/AGGREGATE(로드)가 동일 클래스로 pickle 복원하도록 여기 정의
# ══════════════════════════════════════════════════════════════════════
import os as _os_boot
OUTPUT_BASE = _os_boot.environ.get(
    'VESSL_OUTPUT_DIR', _os_boot.environ.get('OUTPUT_BASE', '/shared'))
try:
    _os_boot.makedirs(OUTPUT_BASE, exist_ok=True)
    _t = _os_boot.path.join(OUTPUT_BASE, '.write_test')
    open(_t,'w').close(); _os_boot.remove(_t)
except Exception:
    OUTPUT_BASE = _os_boot.path.abspath('./results_vessl')
    _os_boot.makedirs(OUTPUT_BASE, exist_ok=True)
RESULTS_DIR = _os_boot.path.join(OUTPUT_BASE, 'results')
_os_boot.makedirs(RESULTS_DIR, exist_ok=True)
print(f'[core73] OUTPUT_BASE = {OUTPUT_BASE}')
print(f'[core73] RESULTS_DIR = {RESULTS_DIR}')

# ══════════════════════════════════════════════════════════════════════
#  Container Stowage Planning — Single Bay, GP-40, 6 PODs
#  PPO Only — Curriculum RL with Weight Constraints (SCIE Research)
#
#  v7 수정사항 (Option C — SCIE 본격 실험):
#   - 레벨별 step 예산 차등 배정 (CONFIG.timesteps_per_level)
#       · Lv1: 150K, Lv2: 300K, Lv3: 500K, Lv4: 1,600K (총 2.25M)
#       · 모든 레벨에서 ~8,000~9,400 에피소드 학습 (균등화)
#   - 보상 구조는 v6 그대로 유지 (R13 Same-Tier Band, R14 Empty Row)
#   - PPO 단독, 100% fill, 6 PODs
# ══════════════════════════════════════════════════════════════════════

# ─────────────────────────────────────────────────────────────────────
# SECTION A: 라이브러리 임포트 & 전역 재현성 확보
# ─────────────────────────────────────────────────────────────────────
import os, sys, random, csv, json, warnings, copy, subprocess, datetime
import numpy as np
import matplotlib
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import matplotlib.gridspec as gridspec
from matplotlib.colors import LinearSegmentedColormap
from scipy import stats as scipy_stats
import gymnasium as gym
from gymnasium import spaces
from stable_baselines3 import PPO                       # ★ PPO만 사용
from stable_baselines3.common.vec_env import DummyVecEnv
from stable_baselines3.common.callbacks import BaseCallback
from dataclasses import dataclass, field
from typing import List, Dict, Tuple, Optional
from collections import defaultdict, Counter

warnings.filterwarnings("ignore")

# ── 전역 시드 고정: 42 ──────────────────────────────────────────────
GLOBAL_SEED = 42

def set_global_seed(seed: int) -> None:
    """모든 난수 생성기를 동일 시드로 고정하여 실험 재현성 확보."""
    random.seed(seed)
    np.random.seed(seed)
    os.environ["PYTHONHASHSEED"] = str(seed)
    try:
        import torch
        torch.manual_seed(seed)
        torch.cuda.manual_seed_all(seed)
        torch.backends.cudnn.deterministic = True
        torch.backends.cudnn.benchmark     = False
    except ImportError:
        pass

set_global_seed(GLOBAL_SEED)
print("✅ [A] Imports & seed fixed (GLOBAL_SEED=42) — PPO only, v6")


# ─────────────────────────────────────────────────────────────────────
# SECTION B: 한글 폰트 설정 (Colab 호환)
# ─────────────────────────────────────────────────────────────────────
def setup_korean_font() -> str:
    """NanumGothic 폰트 설치 및 matplotlib 등록."""
    subprocess.run(["apt-get", "install", "-y", "fonts-nanum"], capture_output=True)
    subprocess.run(["fc-cache", "-fv"], capture_output=True)
    import matplotlib.font_manager as fm
    for font in fm.fontManager.ttflist:
        if "Nanum" in font.name and "Gothic" in font.name:
            plt.rcParams["font.family"]        = font.name
            plt.rcParams["axes.unicode_minus"] = False
            return font.name
    plt.rcParams["font.family"]        = "DejaVu Sans"
    plt.rcParams["axes.unicode_minus"] = False
    return "DejaVu Sans"

_active_font = setup_korean_font()
print(f"✅ [B] Korean font active: {_active_font}")


# ─────────────────────────────────────────────────────────────────────
# SECTION C: 적재 환경 — SingleBayStowageEnv (v6 보상 구조)
#
#  ★ POD 6개, 커리큘럼 4단계 (4×4→6×6→8×8→10×10)
#  ★ MAX_ROWS = 10, OBS_DIM = 79
#
#  ★★★ 보상 구조 v6 ★★★
#
#  [핵심 원칙]
#    1. 컨테이너 선적 순서: POD 내림차순(먼 항구→가까운 항구),
#       동일 POD 내 무게 내림차순(무거운 것 먼저)
#    2. 같은 POD 컨테이너는 같은 Tier에 수평으로 우선 배치 (★ R13 신규)
#    3. 모든 Row에 최소 1개 이상 컨테이너 적재 권장 (R14)
#
#  [Step-level 보상]  (★ v15 조정 가중치 / 계층: 필수·안정·효율)
#    R1   유효 적재 보너스              +1.0          [필수]
#    R2   스택 가득참 패널티            −10.0         [필수]
#    R3   오버스토우 패널티             −15.0         [필수]
#    R4   올바른 하역 순서 보너스       +2.0          [효율] ★ 3.0→2.0
#    R5   중량 균형 보너스              +6.0 × (1−CV) [안정] ★ 3.0→6.0
#    R6   COG 편차 패널티              −8.0 × 편차비율 [안정] ★ −4.0→−8.0
#    R9   열 무게 제한 초과 패널티      −5.0          [필수]
#    R10  같은 층 같은 POD 보너스       +2.67         [효율] ★ 4.0→2.67
#    R11  무게 역전 패널티              −6.0          [필수]
#    R13  ★ Same-Tier Band 보너스      +1.33 (target tier 일치) [효율] ★ 2.0→1.33
#                                      −1.0 × (위로 벗어난 tier 수)  ★ −1.5→−1.0
#    R15  수직 동일-POD 인접 패널티     −2.67         [효율] ★ −4.0→−2.67
#
#  [Terminal 보상]
#    R7   에피소드 완료 보너스          +80.0 × VPR   [필수]
#    R8   POD 수평 밴드 품질 보너스     +1.33         [효율] ★ 2.0→1.33
#    R12  열별 하역순서 완벽도 보너스   +2.0          [효율] ★ 3.0→2.0
#    R14  빈 Row 패널티                 −2.0 × (빈 Row 수) [효율] ★ −3.0→−2.0
# ─────────────────────────────────────────────────────────────────────

# ── POD 상수 (6개 항구) ──────────────────────────────────────────────
N_POD      = 6
POD_NAMES  = {
    1: "Busan",      # 부산 (가장 가까운 항)
    2: "Shanghai",   # 상하이
    3: "Ningbo",     # 닝보
    4: "Singapore",  # 싱가포르
    5: "Colombo",    # 콜롬보
    6: "Rotterdam",  # 로테르담 (가장 먼 항)
}
MAX_ROWS   = 10

# ── 무게 제약조건 상수 ───────────────────────────────────────────────
MIN_WT     = 10.0
MAX_WT     = 20.0
MAX_COL_WT = 145.0

# OBS_DIM = MAX_ROWS*6 + (1+N_POD) + 6 + N_POD = 60 + 7 + 6 + 6 = 79
OBS_DIM    = MAX_ROWS * 6 + (1 + N_POD) + 6 + N_POD


@dataclass
class StowageMetrics:
    """에피소드 레벨 성능 지표 (R13 의미만 변경, 나머지는 동일)."""
    n_total      : int   = 0
    n_valid      : int   = 0
    n_invalid    : int   = 0
    n_overstow   : int   = 0
    n_col_wt_viol: int   = 0
    n_empty_rows : int   = 0
    row_weights  : List[float] = field(default_factory=list)
    pod_per_row  : List[List[int]] = field(default_factory=list)
    # ── 보상 항목별 누적값 (R1~R14, R13 의미 변경) ──
    reward_components : Dict[str, float] = field(default_factory=lambda: {
        "R1_valid": 0.0,
        "R2_stack_full": 0.0,
        "R3_overstow": 0.0,
        "R4_order": 0.0,
        "R5_weight_bal": 0.0,
        "R6_cog": 0.0,
        "R7_completion": 0.0,
        "R8_pod_band": 0.0,
        "R9_col_wt": 0.0,
        "R10_tier_match": 0.0,
        "R11_wt_inversion": 0.0,
        "R12_col_order": 0.0,
        "R13_tier_band": 0.0,           # ★ v6: 의미 변경 (Same-Tier Band)
        "R14_empty_row": 0.0,
        "R15_vstack_pod": 0.0,          # ★ v14: 수직 동일 POD 인접 적재 패널티
    })

    @property
    def vpr(self) -> float:
        return self.n_valid / max(self.n_total, 1)

    @property
    def osr(self) -> float:
        return self.n_overstow / max(self.n_total, 1)

    @property
    def wbi(self) -> float:
        wts = [w for w in self.row_weights if w > 0]
        if len(wts) < 2: return 1.0
        cv = float(np.std(wts)) / (float(np.mean(wts)) + 1e-9)
        return float(np.clip(1.0 - cv, 0.0, 1.0))

    @property
    def psr(self) -> float:
        occupied = [pods for pods in self.pod_per_row if pods]
        if not occupied: return 0.0
        pure = sum(1 for pods in occupied if len(set(pods)) == 1)
        return pure / len(occupied)

    @property
    def cwvr(self) -> float:
        return self.n_col_wt_viol / max(self.n_total, 1)


class SingleBayStowageEnv(gym.Env):
    """
    단일 베이 컨테이너 적재계획 환경 (v6 보상 구조).
    ★ 6개 POD, 커리큘럼 4단계 (4×4 → 10×10), 100% 적재율.
    ★ 컨테이너 선적 순서: POD 내림차순 → 동일 POD 무게 내림차순
    ★ R13 Same-Tier Band 보상으로 같은 POD가 같은 Tier에 수평 적재되도록 유도
    """
    metadata = {"render_modes": []}

    def __init__(self, curriculum_level: int = 1, config: dict = None):
        super().__init__()
        self.cfg   = config if config else CONFIG
        self.level = curriculum_level
        lv         = self.cfg["levels"][self.level]

        self.n_rows       = lv["n_rows"]
        self.n_tiers      = lv["n_tiers"]
        self.n_containers = lv["n_containers"]
        self.n_pods       = N_POD

        # ── 행동 공간: MAX_ROWS(10)로 고정 → 전이학습 호환 ──
        self.action_space = spaces.Discrete(MAX_ROWS)

        # ── 관측 공간: 고정 차원 (OBS_DIM = 79) ──
        self.observation_space = spaces.Box(
            low=-2.0, high=2.0, shape=(OBS_DIM,), dtype=np.float32)

        # ── 내부 상태 ──
        self.wt_grid  : np.ndarray = None
        self.pod_grid : np.ndarray = None
        self.stack_h  : np.ndarray = None
        self.ctns_wt  : np.ndarray = None
        self.ctns_pod : np.ndarray = None
        self.step_idx : int        = 0
        self.metrics  : StowageMetrics = None

    def reset(self, seed: int = None, options: dict = None):
        super().reset(seed=seed)
        rng = np.random.default_rng(seed)

        self.wt_grid  = np.zeros((self.n_rows, self.n_tiers), dtype=np.float32)
        self.pod_grid = np.zeros((self.n_rows, self.n_tiers), dtype=np.int32)
        self.stack_h  = np.zeros(self.n_rows,                 dtype=np.int32)
        self.step_idx = 0
        self.metrics  = StowageMetrics()

        raw_wt  = rng.uniform(MIN_WT, MAX_WT, self.n_containers).astype(np.float32)
        raw_pod = rng.integers(1, self.n_pods + 1, self.n_containers).astype(np.int32)

        # POD 내림차순, 동일 POD 내 무게 내림차순 정렬
        sort_keys = np.lexsort((
            -raw_wt,    # 2차 키: 무게 내림차순
            -raw_pod,   # 1차 키: POD 내림차순
        ))
        self.ctns_wt  = raw_wt[sort_keys]
        self.ctns_pod = raw_pod[sort_keys]

        return self._obs(), {}

    def _get_col_weight(self, row: int) -> float:
        """특정 열(row)에 적재된 컨테이너의 총 무게를 반환."""
        h = int(self.stack_h[row])
        if h == 0:
            return 0.0
        return float(self.wt_grid[row, :h].sum())

    # ─────────────────────────────────────────────────────────────
    # ★ v6 신규 헬퍼: 현재 POD에 대한 'target tier' 계산
    # ─────────────────────────────────────────────────────────────
    def _get_target_tier_for_pod(self, pod: int) -> int:
        """현재 POD에 적합한 target tier 반환.

        규칙:
          (a) 같은 POD가 이미 적재되어 있는 경우:
              - 그 중 가장 높은 tier(=현재 'band' 윗면)를 후보로 함.
              - 그 tier에 빈 자리가 있으면(어떤 row의 stack_h == band_top)
                target = band_top  (수평 확장 권장)
              - 빈 자리가 없으면(전체 row의 stack_h > band_top)
                target = band_top + 1  (정상적으로 한 단 위로 이동)
          (b) 같은 POD가 아직 없는 경우:
              - target = stack_h.min()  (가장 낮은 사용 가능 tier에서 시작)
        """
        # (a) 같은 POD가 적재된 모든 (row, tier) 위치 탐색
        tiers_with_pod: List[int] = []
        for r in range(self.n_rows):
            h = int(self.stack_h[r])
            for t in range(h):
                if int(self.pod_grid[r, t]) == pod:
                    tiers_with_pod.append(t)

        if tiers_with_pod:
            band_top = max(tiers_with_pod)
            # band_top tier에 빈 자리(어떤 row가 정확히 stack_h == band_top)가 있는지
            has_room_on_band = any(int(self.stack_h[r]) == band_top
                                   for r in range(self.n_rows))
            if has_room_on_band:
                return band_top              # 같은 tier에서 수평 확장
            else:
                return band_top + 1          # 다음 tier로 정상 이동
        else:
            # (b) 같은 POD가 아직 없음
            return int(self.stack_h.min())

    def _obs(self) -> np.ndarray:
        """정규화된 관측 벡터 생성 (dim = OBS_DIM = 79)."""
        mid_row = (self.n_rows - 1) / 2.0 + 1e-9

        # ── 스택별 특성 (MAX_ROWS × 6 = 60) ─────────────────────
        stack_feats = np.zeros((MAX_ROWS, 6), dtype=np.float32)
        for r in range(self.n_rows):
            h = int(self.stack_h[r])
            if h == 0:
                stack_feats[r] = [0.0, 0.0, 0.0, 0.0, 0.0, 0.0]
                continue
            fill        = h / self.n_tiers
            wt_norm     = float(self.wt_grid[r, :h].sum()) / (MAX_WT * self.n_tiers)
            mean_pod    = float(self.pod_grid[r, :h].mean()) / self.n_pods
            top_pod     = float(self.pod_grid[r, h - 1])    / self.n_pods
            ov          = sum(1 for t in range(h - 1)
                              if self.pod_grid[r, t] < self.pod_grid[r, t + 1])
            ov_ratio    = ov / max(h - 1, 1)
            col_wt_ratio = self._get_col_weight(r) / MAX_COL_WT
            stack_feats[r] = [fill, wt_norm, mean_pod, top_pod, ov_ratio, col_wt_ratio]

        # ── 현재 컨테이너 특성 [1 + N_POD = 7] ──────────────────
        if self.step_idx < self.n_containers:
            w   = float(self.ctns_wt[self.step_idx])  / MAX_WT
            pod = int(self.ctns_pod[self.step_idx])
            pod_oh = np.eye(self.n_pods, dtype=np.float32)[pod - 1]
            cur_feats = np.array([w, *pod_oh], dtype=np.float32)
        else:
            cur_feats = np.zeros(1 + self.n_pods, dtype=np.float32)

        # ── 글로벌 특성 [6] ──────────────────────────────────────
        total_wt  = float(self.wt_grid.sum()) + 1e-9
        row_wts   = self.wt_grid.sum(axis=1)
        row_wts_nonzero = row_wts[row_wts > 0]

        row_idx  = np.arange(self.n_rows, dtype=np.float32)
        cog      = float(np.dot(row_idx, row_wts)) / total_wt
        cog_norm = (cog - mid_row) / mid_row

        if len(row_wts_nonzero) >= 2:
            wt_cv = float(np.std(row_wts_nonzero)) / (float(np.mean(row_wts_nonzero)) + 1e-9)
        else:
            wt_cv = 0.0

        n_placed = self.step_idx
        total_ov = sum(
            sum(1 for t in range(int(self.stack_h[r]) - 1)
                if self.pod_grid[r, t] < self.pod_grid[r, t + 1])
            for r in range(self.n_rows))

        global_feats = np.array([
            self.step_idx / self.n_containers,
            float(np.clip(cog_norm, -2.0, 2.0)),
            float(np.clip(wt_cv, 0.0, 2.0)),
            self.metrics.n_valid   / max(n_placed, 1),
            total_ov / max(n_placed, 1),
            n_placed / (self.n_rows * self.n_tiers),
        ], dtype=np.float32)

        # ── 잔여 POD 분포 [N_POD = 6] ───────────────────────────
        rem = self.ctns_pod[self.step_idx:]
        if len(rem) > 0:
            counts = np.array([np.sum(rem == p) for p in range(1, self.n_pods + 1)],
                              dtype=np.float32)
            rem_dist = counts / counts.sum()
        else:
            rem_dist = np.zeros(self.n_pods, dtype=np.float32)

        return np.concatenate([
            stack_feats.flatten(),   # 60
            cur_feats,               #  7
            global_feats,            #  6
            rem_dist,                #  6
        ]).astype(np.float32)        # 총 = 79

    def step(self, action) -> Tuple[np.ndarray, float, bool, bool, dict]:
        """컨테이너 적재 실행 및 보상 계산 (v6 보상 구조).

        ★★★ v6 변경점 ★★★
        - R13 (중앙 Row 보너스) 제거
        - R13 (Same-Tier Band 보너스) 신설
        """
        row    = int(action)
        reward = 0.0
        rw     = self.cfg["rw"]

        if self.step_idx >= self.n_containers:
            return self._obs(), 0.0, True, False, {}

        # ── 유효하지 않은 행 선택 (커리큘럼 호환) ──
        if row >= self.n_rows:
            reward += rw["stack_full"] * 2.0
            self.metrics.reward_components["R2_stack_full"] += rw["stack_full"] * 2.0
            self.metrics.n_invalid += 1
            self.step_idx += 1
            terminated = (self.step_idx >= self.n_containers)
            if terminated:
                reward += self._compute_terminal_rewards(rw)
            return self._obs(), float(reward), terminated, False, {
                "valid": self.metrics.n_valid, "invalid": self.metrics.n_invalid,
                "overstow": self.metrics.n_overstow, "step": self.step_idx
            }

        cw  = float(self.ctns_wt[self.step_idx])
        pod = int(self.ctns_pod[self.step_idx])
        h   = int(self.stack_h[row])
        self.metrics.n_total += 1

        # ── R2: 스택 가득참 ──
        if h >= self.n_tiers:
            reward += rw["stack_full"]
            self.metrics.reward_components["R2_stack_full"] += rw["stack_full"]
            self.metrics.n_invalid += 1
            self.step_idx += 1
        else:
            # ── R3 & R4: 오버스토우 / 하역순서 ──
            if h > 0:
                top_pod = int(self.pod_grid[row, h - 1])
                if top_pod < pod:
                    reward += rw["overstow"]
                    self.metrics.reward_components["R3_overstow"] += rw["overstow"]
                    self.metrics.n_overstow += 1
                elif top_pod > pod:
                    reward += rw["order_bonus"]
                    self.metrics.reward_components["R4_order"] += rw["order_bonus"]
                else:
                    reward += rw["order_bonus"] * 0.3
                    self.metrics.reward_components["R4_order"] += rw["order_bonus"] * 0.3
            else:
                reward += rw["order_bonus"] * 0.3
                self.metrics.reward_components["R4_order"] += rw["order_bonus"] * 0.3

            # ── R1: 유효 적재 ──
            reward += rw["valid"]
            self.metrics.reward_components["R1_valid"] += rw["valid"]

            # ─────────────────────────────────────────────────
            # ★ v6: R13 Same-Tier Band 보상 (배치 직전에 계산)
            # 배치 결과 tier(actual_tier) = h, target_tier는 _get_target_tier_for_pod 사용
            # ─────────────────────────────────────────────────
            target_tier = self._get_target_tier_for_pod(pod)
            actual_tier = h

            if actual_tier == target_tier:
                # 정확한 tier에 적재 → 보너스
                r13_val = rw["tier_band_bonus"]
            elif actual_tier > target_tier:
                # 너무 위에 적재 → 패널티 (위로 벗어난 tier 수에 비례)
                r13_val = rw["tier_band_penalty"] * (actual_tier - target_tier)
            else:
                # actual < target: 흔치 않으나 약한 패널티
                r13_val = rw["tier_band_penalty"] * 0.5
            reward += r13_val
            self.metrics.reward_components["R13_tier_band"] += r13_val

            # ── 컨테이너 배치 ──
            self.wt_grid[row, h]  = cw
            self.pod_grid[row, h] = pod
            self.stack_h[row]     = h + 1
            self.metrics.n_valid += 1

            # ── R10: 같은 층(Tier) 같은 POD 매칭 보너스 ──
            tier_match_count = 0
            for other_r in range(self.n_rows):
                if other_r == row:
                    continue
                if int(self.stack_h[other_r]) > h:
                    if int(self.pod_grid[other_r, h]) == pod:
                        tier_match_count += 1
            if tier_match_count > 0:
                r10_val = rw["tier_pod_match"] * min(tier_match_count, 4)  # ★ v14: cap 3→4 (수평 확산 강화)
                reward += r10_val
                self.metrics.reward_components["R10_tier_match"] += r10_val

            # ── R15: 수직 동일 POD 인접 적재 패널티 (★ v14 신규) ──
            #   같은 열에서 바로 아래 칸(h-1)이 동일 POD이면 패널티.
            #   "같은 POD는 같은 층(tier)에 수평 적재" 원칙을 강제하기 위해
            #   세로로 같은 POD를 쌓는 행위(R8, R9 열 세로 적재 문제)를 직접 억제한다.
            #   step 단위 dense 신호로 작동하여 R3/R4/R12의 '단일 POD 세로열' 선호를 상쇄.
            if h > 0:
                below_pod = int(self.pod_grid[row, h - 1])
                if below_pod == pod:
                    # 같은 열에 이미 동일 POD가 몇 칸 연속으로 쌓였는지 카운트 (런 길이)
                    run = 1
                    t = h - 1
                    while t >= 0 and int(self.pod_grid[row, t]) == pod:
                        run += 1
                        t   -= 1
                    # 런이 길수록 강하게 처벌 (선형 누진, 상한 3)
                    r15_val = rw["vstack_pod_penalty"] * min(run - 1, 3)
                    reward += r15_val
                    self.metrics.reward_components["R15_vstack_pod"] += r15_val

            # ── R11: 무게 역전 패널티 ──
            if h > 0:
                below_wt = float(self.wt_grid[row, h - 1])
                if cw > below_wt + 0.5:
                    r11_val = rw["wt_inversion_penalty"]
                    reward += r11_val
                    self.metrics.reward_components["R11_wt_inversion"] += r11_val

            # ── R9: 열 무게 제한 초과 패널티 ──
            col_wt = self._get_col_weight(row)
            if col_wt > MAX_COL_WT:
                r9_val = rw["col_wt_penalty"]
                reward += r9_val
                self.metrics.reward_components["R9_col_wt"] += r9_val
                self.metrics.n_col_wt_viol += 1

            # ── R5: 중량 균형 ──
            occ_wts = np.array([self.wt_grid[r, :int(self.stack_h[r])].sum()
                                for r in range(self.n_rows)], dtype=np.float32)
            used_wts = occ_wts[occ_wts > 0]
            if len(used_wts) >= 2:
                wt_cv = float(np.std(used_wts)) / (float(np.mean(used_wts)) + 1e-9)
                r5_val = rw["weight_balance"] * (1.0 - min(wt_cv, 1.0))
                reward += r5_val
                self.metrics.reward_components["R5_weight_bal"] += r5_val

            # ── R6: COG 패널티 ──
            total_wt = occ_wts.sum() + 1e-9
            cog      = float(np.dot(np.arange(self.n_rows, dtype=np.float32), occ_wts)) / total_wt
            cog_dev  = abs(cog - (self.n_rows - 1) / 2.0) / ((self.n_rows - 1) / 2.0 + 1e-9)
            r6_val   = rw["cog_penalty"] * cog_dev
            reward  += r6_val
            self.metrics.reward_components["R6_cog"] += r6_val

            self.step_idx += 1

        terminated = (self.step_idx >= self.n_containers)

        # ══════════════════════════════════════════════════════
        # ── 종료 보상 (Terminal Rewards) ──
        # ══════════════════════════════════════════════════════
        if terminated:
            reward += self._compute_terminal_rewards(rw)

        return self._obs(), float(reward), terminated, False, {
            "valid": self.metrics.n_valid, "invalid": self.metrics.n_invalid,
            "overstow": self.metrics.n_overstow, "step": self.step_idx,
        }

    def _compute_terminal_rewards(self, rw: dict) -> float:
        """에피소드 종료 시 터미널 보상 계산 (R7, R8, R12, R14)."""
        terminal_reward = 0.0

        # ── R7: 에피소드 완료 보너스 ──
        r7_val = rw["completion"] * self.metrics.vpr
        terminal_reward += r7_val
        self.metrics.reward_components["R7_completion"] += r7_val

        # ── R8: POD 수평 밴드 품질 보너스 ──
        total_band_score = 0.0
        tiers_evaluated  = 0
        for t in range(self.n_tiers):
            pods_at_tier = []
            for r in range(self.n_rows):
                if int(self.stack_h[r]) > t:
                    pods_at_tier.append(int(self.pod_grid[r, t]))
            if len(pods_at_tier) >= 2:
                pod_counts = Counter(pods_at_tier)
                most_common_count = pod_counts.most_common(1)[0][1]
                purity = most_common_count / len(pods_at_tier)
                total_band_score += purity
                tiers_evaluated += 1
        if tiers_evaluated > 0:
            avg_purity = total_band_score / tiers_evaluated
            r8_val = rw["pod_band_quality"] * avg_purity * tiers_evaluated
            terminal_reward += r8_val
            self.metrics.reward_components["R8_pod_band"] += r8_val

        # ── R12: 열별 하역순서 완벽도 보너스 ──
        perfect_cols = 0
        for r in range(self.n_rows):
            h = int(self.stack_h[r])
            if h >= 2:
                is_perfect = True
                for t in range(h - 1):
                    if self.pod_grid[r, t] < self.pod_grid[r, t + 1]:
                        is_perfect = False
                        break
                if is_perfect:
                    perfect_cols += 1
        if perfect_cols > 0:
            r12_val = rw["col_order_bonus"] * perfect_cols
            terminal_reward += r12_val
            self.metrics.reward_components["R12_col_order"] += r12_val

        # ── R14: 빈 Row 패널티 ──
        empty_count = 0
        for r in range(self.n_rows):
            if int(self.stack_h[r]) == 0:
                empty_count += 1
        if empty_count > 0:
            r14_val = rw["empty_row_penalty"] * empty_count
            terminal_reward += r14_val
            self.metrics.reward_components["R14_empty_row"] += r14_val
            self.metrics.n_empty_rows = empty_count

        # ── 메트릭 기록 ──
        self.metrics.row_weights = [
            float(self.wt_grid[r, :int(self.stack_h[r])].sum())
            for r in range(self.n_rows)]
        self.metrics.pod_per_row = [
            list(self.pod_grid[r, :int(self.stack_h[r])])
            for r in range(self.n_rows)]

        return terminal_reward


print("✅ [C] SingleBayStowageEnv defined (v6: Same-Tier Band reward)")
print(f"   Obs dim (fixed): {OBS_DIM}  (MAX_ROWS={MAX_ROWS}×6 + {1+N_POD} + 6 + {N_POD})")
print(f"   PODs ({N_POD}): {POD_NAMES}")
print(f"   Cargo: GP_40 only (weight {MIN_WT:.0f}–{MAX_WT:.0f} MT)")
print(f"   Max column weight: {MAX_COL_WT:.0f} MT")
print(f"   ★ Loading order: POD desc → Weight desc")
print(f"   ★ R13: Same-Tier Band bonus (REPLACES center row bonus)")
print(f"   ★ R14: Empty Row penalty")


# ─────────────────────────────────────────────────────────────────────
# SECTION D: 실험 인프라 — Logger, Callback, CurriculumManager
# ─────────────────────────────────────────────────────────────────────

def make_env(level: int, config: dict, seed: int = 0) -> gym.Env:
    """PPO용 환경 생성."""
    return SingleBayStowageEnv(curriculum_level=level, config=config)


class ExperimentLogger:
    """CSV + JSON 결과 저장 로거."""

    HEADER = ["algorithm","seed","round","level",
              "avg_reward","std_reward","min_reward","max_reward",
              "avg_osr","avg_vpr","avg_wbi","avg_psr","avg_cwvr"]

    def __init__(self, name: str, config: dict):
        ts           = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        self.log_dir = os.path.join(RESULTS_DIR, f"{name}_{ts}")
        os.makedirs(self.log_dir, exist_ok=True)

        cfg_safe = {}
        for k, v in config.items():
            if k == "levels":
                cfg_safe[k] = {str(lk): {sk: sv for sk, sv in lv.items()}
                               for lk, lv in v.items()}
            else:
                cfg_safe[k] = v
        with open(f"{self.log_dir}/config.json", "w", encoding="utf-8") as f:
            json.dump(cfg_safe, f, indent=2, ensure_ascii=False)

        self.csv_path = f"{self.log_dir}/training_log.csv"
        with open(self.csv_path, "w", newline="", encoding="utf-8") as f:
            csv.writer(f).writerow(self.HEADER)
        print(f"✅ Logger → {self.log_dir}")

    def log(self, algo, seed, rnd, lvl, rewards, metrics_list):
        avg_osr  = float(np.mean([m.osr  for m in metrics_list]))
        avg_vpr  = float(np.mean([m.vpr  for m in metrics_list]))
        avg_wbi  = float(np.mean([m.wbi  for m in metrics_list]))
        avg_psr  = float(np.mean([m.psr  for m in metrics_list]))
        avg_cwvr = float(np.mean([m.cwvr for m in metrics_list]))
        with open(self.csv_path, "a", newline="", encoding="utf-8") as f:
            csv.writer(f).writerow([
                algo, seed, rnd, lvl,
                f"{np.mean(rewards):.4f}", f"{np.std(rewards):.4f}",
                f"{np.min(rewards):.4f}",  f"{np.max(rewards):.4f}",
                f"{avg_osr:.4f}", f"{avg_vpr:.4f}",
                f"{avg_wbi:.4f}", f"{avg_psr:.4f}",
                f"{avg_cwvr:.4f}",
            ])

    def save_fig(self, fig, name: str) -> str:
        path = f"{self.log_dir}/{name}"
        fig.savefig(path, dpi=300, bbox_inches="tight", facecolor="white")
        print(f"  💾 Figure saved: {path}")
        return path


class EpisodeMetricsCallback(BaseCallback):
    """에피소드별 보상 및 길이 기록 + ★ v8: 목표 에피소드 도달 시 학습 종료."""

    def __init__(self, target_episodes: Optional[int] = None):
        super().__init__(verbose=0)
        self.ep_rewards  : List[float] = []
        self.ep_lengths  : List[int]   = []
        self.timesteps   : List[int]   = []
        self._cur_r      = 0.0
        self._cur_l      = 0
        # ★ v8: 목표 에피소드 (None 이면 step-based 종료, SB3 기본 동작)
        self.target_episodes = target_episodes

    def _on_step(self) -> bool:
        self._cur_r += float(self.locals["rewards"][0])
        self._cur_l += 1
        if self.locals["dones"][0]:
            self.ep_rewards.append(self._cur_r)
            self.ep_lengths.append(self._cur_l)
            self.timesteps.append(self.num_timesteps)
            self._cur_r = 0.0
            self._cur_l = 0
            # ★ v8: 목표 에피소드 도달 → False 반환하면 model.learn() 종료
            if (self.target_episodes is not None
                and len(self.ep_rewards) >= self.target_episodes):
                return False
        return True

    def smoothed(self, window: int = 20) -> np.ndarray:
        r = np.array(self.ep_rewards)
        if len(r) < window: return r
        kernel = np.ones(window) / window
        return np.convolve(r, kernel, mode="valid")


class CurriculumManager:
    """레벨 1→2→3→4 점진적 난이도 관리."""

    def __init__(self, config: dict):
        self.config    = config
        self.level     = 1
        self.max_level = len(config["levels"])

    def should_level_up(self, avg_r: float) -> bool:
        return avg_r > self.config["level_up_threshold"] and self.level < self.max_level

    def level_up(self):
        if self.level < self.max_level:
            self.level += 1
            lv = self.config["levels"][self.level]
            print(f"  🆙 Level UP → Lv{self.level}: "
                  f"{lv['n_containers']} containers, "
                  f"{lv['n_rows']} rows × {lv['n_tiers']} tiers")

print("✅ [D] Logger, Callback, CurriculumManager defined")


# ─────────────────────────────────────────────────────────────────────
# SECTION E: PPO 모델 구축 & 학습 루프
# ─────────────────────────────────────────────────────────────────────

def _tb_log_dir(config: dict):
    """tensorboard 사용 가능 시에만 로그 경로 반환 (없으면 None → 학습 중단 방지)."""
    try:
        import tensorboard  # noqa: F401
    except Exception:
        return None
    d = os.path.join(RESULTS_DIR, 'tb', config.get('experiment_name', 'ppo'))
    os.makedirs(d, exist_ok=True)
    return d


def _build_ppo_model(env, config: dict, seed: int):
    """PPO 모델 생성."""
    p = config["ppo"]
    return PPO(
        "MlpPolicy", env,
        learning_rate  = p["lr"],
        gamma          = p["gamma"],
        n_steps        = p["n_steps"],
        batch_size     = p["batch_size"],
        n_epochs       = p["n_epochs"],
        clip_range     = p["clip_range"],
        ent_coef       = p["ent_coef"],
        gae_lambda     = p["gae_lambda"],
        policy_kwargs  = {"net_arch": p["net_arch"]},
        seed           = seed,
        tensorboard_log = _tb_log_dir(config),
        verbose        = 0,
        device         = "auto",
    )


def _evaluate_episode_full(model, level: int, config: dict, seed: int
                           ) -> Tuple[float, 'StowageMetrics', dict]:
    """평가 에피소드 1회 실행 + Bay Plan 상태 반환."""
    env = make_env(level=level, config=config)
    obs, _ = env.reset(seed=seed)
    done   = False; total_r = 0.0
    while not done:
        act, _ = model.predict(obs, deterministic=True)
        obs, r, term, trunc, _ = env.step(act)
        done = term or trunc
        total_r += r

    bay_plan = {
        "wt_grid":  env.wt_grid.copy(),
        "pod_grid": env.pod_grid.copy(),
        "stack_h":  env.stack_h.copy(),
        "ctns_wt":  env.ctns_wt.copy(),
        "ctns_pod": env.ctns_pod.copy(),
        "n_rows":   env.n_rows,
        "n_tiers":  env.n_tiers,
    }
    return total_r, env.metrics, bay_plan


def _run_training(
    config    : dict,
    logger    : ExperimentLogger,
    seed      : int,
) -> Dict:
    """PPO 학습 루프."""
    set_global_seed(seed)
    algo_name   = "PPO"
    curriculum  = CurriculumManager(config)
    results     = {"algo": algo_name,
                   "avg_rewards": [], "std_rewards": [], "levels": [],
                   "callbacks": [], "final_metrics": None, "trained_model": None,
                   "per_round_metrics": [],
                   "per_round_bay_plans": [],
                   "phase_info":          []}     # ★ v11 NEW
    model       = None

    # ★ v11 NEW: Lexicographic phase 설정
    lex_spec = config.get("lexicographic", None)
    base_rw  = dict(config["rw"])    # BASE rewards 원본 보존

    for rnd in range(config["total_rounds"]):
        lv     = curriculum.level
        lv_cfg = config["levels"][lv]
        print(f"  [{algo_name} seed={seed}] Round {rnd+1}/{config['total_rounds']} "
              f"Lv{lv}: {lv_cfg['n_containers']}c × {lv_cfg['n_rows']}r×{lv_cfg['n_tiers']}t")

        # ★ v11 NEW: Phase 결정 + 보상 가중치 적용 (round 번호 기반 — robust)
        if lex_spec is not None:
            current_round = rnd + 1   # 1-indexed
            if current_round in lex_spec["phase1_rounds"]:
                phase, mode = 1, lex_spec["phase1_mode"]
            elif current_round in lex_spec["phase2_rounds"]:
                phase, mode = 2, lex_spec["phase2_mode"]
            else:
                phase, mode = 2, "BASE"
            config["rw"] = make_phase_weights(base_rw, mode)
            n_active = sum(1 for v in config["rw"].values() if v != 0)
            print(f"    🎯 Phase {phase} ({mode}): {n_active}/{len(base_rw)} 보상 활성")
        else:
            phase, mode = 0, "BASE"
            config["rw"] = base_rw

        results["phase_info"].append({
            "round": rnd + 1, "level": lv, "phase": phase, "mode": mode,
            "n_active_rewards": sum(1 for v in config["rw"].values() if v != 0),
            "rw_snapshot": dict(config["rw"]),
        })

        _lv = lv
        def _make(l=_lv): return make_env(l, config, seed)
        train_env = DummyVecEnv([_make])

        if model is None:
            model = _build_ppo_model(train_env, config, seed)
            print(f"    ✅ {algo_name} Model created | obs:{OBS_DIM} | act:{train_env.action_space}")
        else:
            model.set_env(train_env)
            print(f"    🔄 Transfer learning: env replaced, weights retained")

        # ★ v8: 레벨별 "에피소드" 예산 (스텝이 아닌 에피소드로 학습 종료)
        eps_per_level = config.get("episodes_per_level")
        if eps_per_level and lv in eps_per_level:
            target_eps = eps_per_level[lv]
        else:
            # 폴백: timesteps_per_round → 에피소드로 환산 (legacy 호환)
            target_eps = (config.get("timesteps_per_round", 200_000)
                          // lv_cfg["n_containers"])

        # SB3 model.learn(total_timesteps) 안전 상한 (callback이 먼저 종료)
        sf      = config.get("max_steps_safety_factor", 2.0)
        max_ts  = int(target_eps * lv_cfg["n_containers"] * sf)
        print(f"    💡 Episode budget for Lv{lv}: {target_eps:,} episodes "
              f"(safety cap: {max_ts:,} steps, ×{sf})")

        import time
        _t0 = time.time()
        cb  = EpisodeMetricsCallback(target_episodes=target_eps)
        model.learn(total_timesteps=max_ts, callback=cb,
                    reset_num_timesteps=False)
        train_sec    = time.time() - _t0
        actual_eps   = len(cb.ep_rewards)
        actual_steps = cb.timesteps[-1] if cb.timesteps else 0
        results["callbacks"].append(cb)
        print(f"    ⏱  Trained {actual_eps:,} episodes "
              f"({actual_steps:,} steps, {train_sec/60:.1f} min)")

        n_eval      = config["eval_episodes"]
        ep_rewards  = []
        ep_metrics  = []
        last_bay_plan = None
        for ep in range(n_eval):
            r, m, bp = _evaluate_episode_full(model, lv, config, seed + ep * 17)
            ep_rewards.append(r)
            ep_metrics.append(m)
            last_bay_plan = bp

        results["per_round_bay_plans"].append(last_bay_plan)

        avg_r = float(np.mean(ep_rewards))
        std_r = float(np.std(ep_rewards))
        avg_empty = float(np.mean([m.n_empty_rows for m in ep_metrics]))
        print(f"    📊 Eval: reward={avg_r:8.2f}±{std_r:6.2f} | "
              f"OSR={np.mean([m.osr for m in ep_metrics]):.3f} | "
              f"VPR={np.mean([m.vpr for m in ep_metrics]):.3f} | "
              f"WBI={np.mean([m.wbi for m in ep_metrics]):.3f} | "
              f"PSR={np.mean([m.psr for m in ep_metrics]):.3f} | "
              f"CWVR={np.mean([m.cwvr for m in ep_metrics]):.3f} | "
              f"EmptyR={avg_empty:.1f}")

        results["avg_rewards"].append(avg_r)
        results["std_rewards"].append(std_r)
        results["levels"].append(lv)
        results["per_round_metrics"].append({
            "round": rnd + 1, "level": lv,
            "avg_reward": avg_r, "std_reward": std_r,
            "avg_osr":  float(np.mean([m.osr  for m in ep_metrics])),
            "avg_vpr":  float(np.mean([m.vpr  for m in ep_metrics])),
            "avg_wbi":  float(np.mean([m.wbi  for m in ep_metrics])),
            "avg_psr":  float(np.mean([m.psr  for m in ep_metrics])),
            "avg_cwvr": float(np.mean([m.cwvr for m in ep_metrics])),
            "avg_empty_rows": avg_empty,
            # ★ v8: 학습 통계 추가
            "actual_episodes": actual_eps,
            "actual_steps":    actual_steps,
            "train_seconds":   train_sec,
        })
        logger.log(algo_name, seed, rnd + 1, lv, ep_rewards, ep_metrics)
        results["final_metrics"] = ep_metrics

        # ════════ ★ v8: 라운드 종료 요약 블록 ════════
        bar = "═" * 72
        print()
        print(f"    {bar}")
        print(f"    🏁 ROUND {rnd+1}/{config['total_rounds']} 완료  "
              f"│  Lv{lv}: {lv_cfg['n_containers']} 컨테이너  "
              f"({lv_cfg['n_rows']}×{lv_cfg['n_tiers']})")
        print(f"    {bar}")
        print(f"      📚 학습: {actual_eps:>6,} episodes  │  "
              f"{actual_steps:>9,} steps  │  {train_sec/60:>5.1f} min")
        print(f"      🧪 평가: reward = {avg_r:+8.2f} ± {std_r:6.2f}  "
              f"(over {n_eval} eval episodes)")
        m_osr  = float(np.mean([m.osr  for m in ep_metrics]))
        m_vpr  = float(np.mean([m.vpr  for m in ep_metrics]))
        m_wbi  = float(np.mean([m.wbi  for m in ep_metrics]))
        m_psr  = float(np.mean([m.psr  for m in ep_metrics]))
        m_cwvr = float(np.mean([m.cwvr for m in ep_metrics]))
        print(f"      📊 지표: OSR={m_osr:.3f}  VPR={m_vpr:.3f}  "
              f"WBI={m_wbi:.3f}  PSR={m_psr:.3f}  "
              f"CWVR={m_cwvr:.3f}  EmptyR={avg_empty:.1f}")
        cum_str = "  →  ".join(f"R{i+1}:{r:+.1f}"
                                for i, r in enumerate(results["avg_rewards"]))
        print(f"      📈 누적: {cum_str}")
        print(f"    {bar}")
        print()
        # ═══════════════════════════════════════════════

        if curriculum.should_level_up(avg_r):
            curriculum.level_up()

    results["trained_model"] = model
    return results


print("✅ [E] PPO training functions defined")


# ─────────────────────────────────────────────────────────────────────
# SECTION F: 출판 품질 시각화 함수 (PPO 단독)
# ─────────────────────────────────────────────────────────────────────

CLR  = {"PPO": "#2166AC"}
FILL = {"PPO": "#74ADD1"}

def _scie_style():
    """출판 품질 matplotlib 기본값 적용."""
    plt.rcParams.update({
        "font.size": 10, "axes.titlesize": 11, "axes.labelsize": 10,
        "xtick.labelsize": 9, "ytick.labelsize": 9, "legend.fontsize": 9,
        "lines.linewidth": 1.8, "lines.markersize": 6,
        "axes.grid": True, "grid.alpha": 0.3,
        "figure.dpi": 100, "savefig.dpi": 300,
        "axes.spines.top": False, "axes.spines.right": False,
    })
    plt.rcParams["font.family"]        = _active_font
    plt.rcParams["axes.unicode_minus"] = False


# ── Figure 1: 학습 곡선 (X축 = Round) ──────────────────────────────
def plot_learning_curves_by_round(
    ppo_res: Dict,
    logger: ExperimentLogger, config: dict
) -> None:
    """Fig.1: 커리큘럼 라운드별 학습 성능."""
    _scie_style()
    fig, ax = plt.subplots(figsize=(10, 5))
    fig.suptitle(
        "Fig. 1  PPO — Curriculum Learning (X-axis: Round)\n"
        f"(v6: Same-Tier Band reward | Single-Bay, GP-40, {N_POD} PODs, "
        f"Weight {MIN_WT:.0f}–{MAX_WT:.0f}MT, Max Col {MAX_COL_WT:.0f}MT, Fill 100%)",
        fontsize=11, fontweight="bold", y=1.02)

    x = np.arange(1, config["total_rounds"] + 1)

    mu = np.array(ppo_res["avg_rewards"])
    sd = np.array(ppo_res["std_rewards"])
    ax.plot(x, mu, "o-", color=CLR["PPO"], label="PPO (mean)", lw=2, ms=7, zorder=3)
    ax.fill_between(x, mu - sd, mu + sd, alpha=0.2, color=FILL["PPO"], label="PPO ±1σ")

    ax.axhline(0, color="gray", ls="--", lw=1, alpha=0.5)

    levels = ppo_res["levels"]
    for i in range(len(levels) - 1):
        if levels[i] != levels[i + 1]:
            ax.axvline(x=x[i + 1] - 0.5, color="orange", ls=":", lw=1.5, alpha=0.8)
    for i, lv in enumerate(levels):
        ax.text(x[i], ax.get_ylim()[0], f"Lv{lv}", ha="center", va="top",
                fontsize=7.5, color="gray")

    ax.set_xlabel("Curriculum Round")
    ax.set_ylabel("Average Episode Reward")
    ax.set_title("PPO — Reward per Round")
    ax.legend(loc="upper left")
    ax.set_xticks(x)

    plt.tight_layout()
    logger.save_fig(fig, "fig1_learning_curves_by_round.png")
    plt.show(); plt.close(fig)


# ── Figure 2: 학습 곡선 (X축 = Step) ───────────────────────────────
def plot_learning_curves_by_step(
    ppo_res: Dict,
    logger: ExperimentLogger, config: dict
) -> None:
    """Fig.2: 학습 스텝별 에피소드 보상."""
    _scie_style()
    fig, ax = plt.subplots(figsize=(10, 5))
    fig.suptitle(
        "Fig. 2  PPO — Learning Curve (X-axis: Training Step)\n"
        "(Smoothed episode reward, window=20 | v6 reward structure)",
        fontsize=11, fontweight="bold", y=1.02)

    all_rewards = []
    all_steps   = []
    for cb in ppo_res["callbacks"]:
        all_rewards.extend(cb.ep_rewards)
        all_steps.extend(cb.timesteps)

    r = np.array(all_rewards)
    s = np.array(all_steps)
    window = min(20, len(r))
    if window > 1:
        kernel = np.ones(window) / window
        r_smooth = np.convolve(r, kernel, mode="valid")
        s_smooth = s[window-1:]
    else:
        r_smooth = r
        s_smooth = s

    ax.plot(s_smooth, r_smooth, color=CLR["PPO"], lw=1.2, alpha=0.8,
            label="PPO (smoothed)")

    ax.set_xlabel("Training Step")
    ax.set_ylabel("Episode Reward")
    ax.set_title("PPO — Reward vs Step")
    ax.legend(loc="upper left")

    plt.tight_layout()
    logger.save_fig(fig, "fig2_learning_curves_by_step.png")
    plt.show(); plt.close(fig)


# ── Figure 3: 라운드별 바 차트 ─────────────────────────────────────
def plot_round_comparison(
    ppo_res: Dict,
    logger: ExperimentLogger, config: dict
) -> None:
    """Fig.3: PPO 라운드별 평균 보상 바차트."""
    _scie_style()
    n_rounds = config["total_rounds"]
    x = np.arange(1, n_rounds + 1)
    w = 0.55

    fig, ax = plt.subplots(figsize=(10, 5))
    fig.suptitle("Fig. 3  PPO Round-wise Performance (v6)",
                 fontsize=11, fontweight="bold")

    means = np.array(ppo_res["avg_rewards"])
    stds  = np.array(ppo_res["std_rewards"])
    bars = ax.bar(x, means, w, yerr=stds, label="PPO",
                  color=CLR["PPO"], alpha=0.85, edgecolor="white",
                  capsize=3, error_kw={"lw": 1})
    for bar in bars:
        h = bar.get_height()
        ax.text(bar.get_x() + bar.get_width()/2., h + (1 if h >= 0 else -3),
                f'{h:.1f}', ha='center', va='bottom' if h >= 0 else 'top',
                fontsize=7)

    levels = ppo_res["levels"]
    for i, lv in enumerate(levels):
        ax.text(x[i], ax.get_ylim()[0] - 3, f"Lv{lv}",
                ha="center", fontsize=8, color="gray", fontstyle="italic")

    ax.set_xlabel("Curriculum Round")
    ax.set_ylabel("Average Episode Reward")
    ax.set_xticks(x)
    ax.legend()
    ax.axhline(0, color="gray", ls="--", lw=0.8, alpha=0.5)

    plt.tight_layout()
    logger.save_fig(fig, "fig3_round_comparison_bar.png")
    plt.show(); plt.close(fig)


# ── Figure 4: 메트릭 레이더 ────────────────────────────────────────
def plot_metrics_radar(
    ppo_res: Dict,
    logger: ExperimentLogger
) -> None:
    """Fig.4: PPO 최종 성능 메트릭 레이더 차트."""
    _scie_style()
    categories = ['AER\n(norm)', 'VPR', 'WBI', 'PSR', '1-OSR', '1-CWVR']

    max_abs_r = max(abs(ppo_res["avg_rewards"][-1]), 1)

    angles = np.linspace(0, 2 * np.pi, len(categories), endpoint=False).tolist()
    angles += angles[:1]

    fig, ax = plt.subplots(figsize=(7, 7), subplot_kw=dict(polar=True))
    fig.suptitle("Fig. 4  PPO Performance Radar (v6)", fontsize=11, fontweight="bold")

    vals = [max(0, ppo_res["avg_rewards"][-1] / max_abs_r)]
    fm = ppo_res["final_metrics"]
    vals.extend([
        float(np.mean([m.vpr for m in fm])),
        float(np.mean([m.wbi for m in fm])),
        float(np.mean([m.psr for m in fm])),
        1 - float(np.mean([m.osr for m in fm])),
        1 - float(np.mean([m.cwvr for m in fm])),
    ])
    vals += vals[:1]
    ax.fill(angles, vals, color=CLR["PPO"], alpha=0.20)
    ax.plot(angles, vals, "o-", color=CLR["PPO"], lw=2, label="PPO")

    ax.set_thetagrids(np.degrees(angles[:-1]), categories)
    ax.set_ylim(0, 1.1)
    ax.legend(loc="upper right", bbox_to_anchor=(1.3, 1.1))

    plt.tight_layout()
    logger.save_fig(fig, "fig4_metrics_radar.png")
    plt.show(); plt.close(fig)


# ── Figure 5: Bay Plan 시각화 — PPO ────────────────────────────────
def plot_bay_plan(model, level: int, config: dict, logger: ExperimentLogger,
                  algo_name: str = "PPO") -> None:
    """Fig.5: Bay Plan 시각화 — POD 분포(좌) & 중량 분포(우)."""
    _scie_style()
    env = make_env(level, config, seed=42)
    obs, _ = env.reset(seed=42)
    done = False
    while not done:
        act, _ = model.predict(obs, deterministic=True)
        obs, _, term, trunc, _ = env.step(act)
        done = term or trunc

    n_rows   = env.n_rows
    n_tiers  = env.n_tiers
    wt_grid  = env.wt_grid
    pod_grid = env.pod_grid
    stack_h  = env.stack_h

    pod_display = np.full((n_tiers, n_rows), np.nan)
    wt_display  = np.zeros((n_tiers, n_rows))
    for r in range(n_rows):
        h = int(stack_h[r])
        for t in range(h):
            pod_display[t, r] = pod_grid[r, t]
            wt_display [t, r] = wt_grid [r, t]

    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(14, max(4, n_tiers * 0.55 + 2)))
    fig.suptitle(
        f"Fig. 5  Bay Plan — {algo_name} (Lv{level}, v6)\n"
        f"Left: POD distribution | Right: Weight (MT) | "
        f"Max Col WT: {MAX_COL_WT:.0f} MT",
        fontsize=11, fontweight="bold", y=1.02)

    pod_colors = ["#DDDDDD", "#4393C3", "#F4A582", "#D6604D", "#92C5DE", "#2166AC", "#B2182B"]
    pod_cmap = LinearSegmentedColormap.from_list("pod6", pod_colors[:N_POD+1], N=N_POD+1)
    im1 = ax1.imshow(pod_display, cmap=pod_cmap, vmin=0, vmax=N_POD,
                     origin="lower", aspect="auto", interpolation="nearest")
    for t in range(n_tiers):
        for r in range(n_rows):
            if not np.isnan(pod_display[t, r]):
                p = int(pod_display[t, r])
                ax1.text(r, t, POD_NAMES.get(p, "?")[:3], ha="center", va="center",
                         fontsize=6, fontweight="bold",
                         color="white" if p >= 3 else "black")
    ax1.set_xlabel("Row (Column)")
    ax1.set_ylabel("Tier (bottom → top)")
    ax1.set_title(f"POD Distribution ({N_POD} ports)")
    ax1.set_xticks(range(n_rows)); ax1.set_xticklabels([f"R{r}" for r in range(n_rows)])
    ax1.set_yticks(range(n_tiers)); ax1.set_yticklabels([f"T{t}" for t in range(n_tiers)])
    plt.colorbar(im1, ax=ax1, label="POD ID", ticks=list(range(N_POD+1)))

    for r in range(n_rows):
        h = int(stack_h[r])
        for t in range(h - 1):
            if pod_grid[r, t] < pod_grid[r, t + 1]:
                ax1.add_patch(mpatches.FancyBboxPatch(
                    (r - 0.45, t - 0.45), 0.9, 0.9,
                    boxstyle="round,pad=0.05", linewidth=2,
                    edgecolor="red", facecolor="none"))

    im2 = ax2.imshow(wt_display, cmap="YlOrRd", vmin=0, vmax=MAX_WT,
                     origin="lower", aspect="auto", interpolation="nearest")
    for r in range(n_rows):
        h = int(stack_h[r])
        for t in range(h):
            ax2.text(r, t, f"{wt_grid[r,t]:.1f}", ha="center", va="center",
                     fontsize=5.5, color="black")

    for r in range(n_rows):
        col_wt = float(wt_grid[r, :int(stack_h[r])].sum())
        color = "red" if col_wt > MAX_COL_WT else "green"
        ax2.text(r, n_tiers - 0.3, f"Σ{col_wt:.0f}", ha="center", va="bottom",
                 fontsize=6, fontweight="bold", color=color)

    m = env.metrics
    ax2.set_xlabel("Row (Column)")
    ax2.set_ylabel("Tier (bottom → top)")
    ax2.set_title(f"Weight (MT) | VPR={m.vpr*100:.1f}% | OSR={m.osr:.3f}")
    ax2.set_xticks(range(n_rows)); ax2.set_xticklabels([f"R{r}" for r in range(n_rows)])
    ax2.set_yticks(range(n_tiers)); ax2.set_yticklabels([f"T{t}" for t in range(n_tiers)])
    plt.colorbar(im2, ax=ax2, label="Weight (MT)")

    plt.tight_layout()
    logger.save_fig(fig, f"fig5_bay_plan_{algo_name}.png")
    plt.show(); plt.close(fig)

    print(f"\n  {'─'*50}")
    print(f"  Bay Plan Summary — {algo_name} (Lv{level})")
    print(f"  {'─'*50}")
    print(f"  Valid / Invalid        : {m.n_valid} / {m.n_invalid}")
    print(f"  Overstow count         : {m.n_overstow}  (OSR={m.osr:.3f})")
    print(f"  Col Weight Violations  : {m.n_col_wt_viol}  (CWVR={m.cwvr:.3f})")
    print(f"  Empty Rows             : {m.n_empty_rows}")
    print(f"  Valid Placement Rate   : {m.vpr:.3f}")
    print(f"  Weight Balance Index   : {m.wbi:.3f}")
    print(f"  POD Segregation Rate   : {m.psr:.3f}")
    print(f"  {'─'*50}")


# ── 결과 요약 테이블 — PPO ──────────────────────────────────────────
def print_round_summary_table(ppo_res: Dict, config: dict):
    """PPO 라운드별 결과 요약 테이블."""
    print("\n" + "═" * 90)
    print("  TABLE: PPO — Round-wise Performance Summary (v6)")
    print("═" * 90)
    print(f"  {'Round':<6} {'Level':<6} {'Reward':>10} {'±Std':>8} "
          f"{'OSR':>7} {'VPR':>7} {'WBI':>7} {'PSR':>7} {'CWVR':>7} {'EmptyR':>7}")
    print("  " + "─" * 86)

    for rnd in range(config["total_rounds"]):
        pm = ppo_res["per_round_metrics"][rnd]
        print(f"  {rnd+1:<6} Lv{pm['level']:<4} "
              f"{pm['avg_reward']:>+9.2f} {pm['std_reward']:>7.2f}  "
              f"{pm['avg_osr']:>6.3f} {pm['avg_vpr']:>6.3f} "
              f"{pm['avg_wbi']:>6.3f} {pm['avg_psr']:>6.3f} "
              f"{pm.get('avg_cwvr', 0):>6.3f} {pm.get('avg_empty_rows', 0):>6.1f}")

    print("═" * 90)

    # ── 최종 라운드 성능 ──
    print(f"\n  FINAL PERFORMANCE (Last Round)")
    print("  " + "─" * 60)
    print(f"  {'Metric':<40} {'PPO':>15}")
    print("  " + "─" * 60)

    fm = ppo_res["final_metrics"]
    metrics = [
        ("Avg. Episode Reward",            ppo_res['avg_rewards'][-1]),
        ("Overstow Rate (OSR)",            float(np.mean([m.osr for m in fm]))),
        ("Valid Placement Rate (VPR)",     float(np.mean([m.vpr for m in fm]))),
        ("Weight Balance Index (WBI)",     float(np.mean([m.wbi for m in fm]))),
        ("POD Segregation Rate (PSR)",     float(np.mean([m.psr for m in fm]))),
        ("Col Weight Viol Rate (CWVR)",    float(np.mean([m.cwvr for m in fm]))),
        ("Empty Rows (avg)",               float(np.mean([m.n_empty_rows for m in fm]))),
    ]
    for label, pv in metrics:
        print(f"  {label:<40} {pv:>+12.3f}")
    print("═" * 90)


# ── Excel 저장 — PPO ───────────────────────────────────────────────
def save_results_to_excel(
    ppo_res: Dict, config: dict,
    seed: int, logger: ExperimentLogger
) -> str:
    """결과를 Excel 파일로 저장."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill_ppo = PatternFill("solid", fgColor="2166AC")
    header_fill_gen = PatternFill("solid", fgColor="404040")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'))

    # ══════ Sheet 1: 라운드별 결과 ══════
    ws1 = wb.active
    ws1.title = "Round Summary"
    headers = ["Round", "Algo", "Level", "Reward", "Std", "OSR", "VPR",
               "WBI", "PSR", "CWVR", "Empty Rows"]
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill_gen
        cell.alignment = center_align

    row_num = 2
    for rnd in range(config["total_rounds"]):
        pm = ppo_res["per_round_metrics"][rnd]
        ws1.cell(row=row_num, column=1, value=rnd+1)
        c = ws1.cell(row=row_num, column=2, value="PPO")
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = header_fill_ppo
        ws1.cell(row=row_num, column=3, value=pm["level"])
        ws1.cell(row=row_num, column=4, value=round(pm["avg_reward"], 2))
        ws1.cell(row=row_num, column=5, value=round(pm["std_reward"], 2))
        ws1.cell(row=row_num, column=6, value=round(pm["avg_osr"], 4))
        ws1.cell(row=row_num, column=7, value=round(pm["avg_vpr"], 4))
        ws1.cell(row=row_num, column=8, value=round(pm["avg_wbi"], 4))
        ws1.cell(row=row_num, column=9, value=round(pm["avg_psr"], 4))
        ws1.cell(row=row_num, column=10, value=round(pm.get("avg_cwvr", 0), 4))
        ws1.cell(row=row_num, column=11, value=round(pm.get("avg_empty_rows", 0), 1))
        row_num += 1

    for col in ws1.columns:
        max_len = max(len(str(cell.value or "")) for cell in col) + 2
        ws1.column_dimensions[col[0].column_letter].width = max(max_len, 12)

    # ══════ Sheet 2~: Bay Plan per round ══════
    pod_fills = {
        0: PatternFill("solid", fgColor="F2F2F2"),
        1: PatternFill("solid", fgColor="4393C3"),
        2: PatternFill("solid", fgColor="F4A582"),
        3: PatternFill("solid", fgColor="D6604D"),
        4: PatternFill("solid", fgColor="92C5DE"),
        5: PatternFill("solid", fgColor="2166AC"),
        6: PatternFill("solid", fgColor="B2182B"),
    }
    white_font = Font(bold=True, color="FFFFFF", size=9)
    black_font = Font(bold=True, color="000000", size=9)

    algo_name = "PPO"
    for rnd_idx, bp in enumerate(ppo_res.get("per_round_bay_plans", [])):
        if bp is None:
            continue
        rnd_num = rnd_idx + 1
        lv = ppo_res["levels"][rnd_idx]
        ws_bp = wb.create_sheet(f"{algo_name}_R{rnd_num}_Lv{lv}")

        n_rows_bp = bp["n_rows"]
        n_tiers_bp = bp["n_tiers"]
        wt_grid = bp["wt_grid"]
        pod_grid = bp["pod_grid"]
        stack_h  = bp["stack_h"]

        # 제목
        ws_bp.merge_cells(start_row=1, start_column=1,
                          end_row=1, end_column=n_rows_bp + 2)
        title_cell = ws_bp.cell(row=1, column=1,
            value=f"{algo_name} Round {rnd_num} (Lv{lv}) — {n_rows_bp}R×{n_tiers_bp}T  [v6]")
        title_cell.font = Font(bold=True, size=12)

        # POD 분포
        pod_start = 3
        ws_bp.cell(row=pod_start, column=1, value="【POD 분포】").font = Font(bold=True, size=11)
        pod_start += 1
        ws_bp.cell(row=pod_start, column=1, value="Tier\\Row")
        for r in range(n_rows_bp):
            cell = ws_bp.cell(row=pod_start, column=r+2, value=f"R{r}")
            cell.font = header_font
            cell.fill = header_fill_ppo
            cell.alignment = center_align

        for t_d in range(n_tiers_bp):
            t_a = n_tiers_bp - 1 - t_d
            dr = pod_start + 1 + t_d
            ws_bp.cell(row=dr, column=1, value=f"T{t_a}").font = Font(bold=True)
            for r in range(n_rows_bp):
                cell = ws_bp.cell(row=dr, column=r+2)
                cell.alignment = center_align
                cell.border = thin_border
                h = int(stack_h[r])
                if t_a < h:
                    p = int(pod_grid[r, t_a])
                    cell.value = f"{POD_NAMES.get(p,'?')[:3]}({p})"
                    cell.fill = pod_fills.get(p, pod_fills[0])
                    cell.font = white_font if p >= 3 else black_font
                else:
                    cell.value = "—"
                    cell.fill = pod_fills[0]

        # 열 합계
        sum_row = pod_start + 1 + n_tiers_bp
        ws_bp.cell(row=sum_row, column=1, value="Col ΣWT").font = Font(bold=True)
        for r in range(n_rows_bp):
            h = int(stack_h[r])
            col_wt = float(wt_grid[r, :h].sum())
            cell = ws_bp.cell(row=sum_row, column=r+2, value=round(col_wt, 1))
            cell.alignment = center_align
            cell.font = Font(bold=True,
                color="FF0000" if col_wt > MAX_COL_WT else "008000")

        # Weight 분포
        wt_start = sum_row + 2
        ws_bp.cell(row=wt_start, column=1, value="【Weight (MT)】").font = Font(bold=True, size=11)
        wt_start += 1
        ws_bp.cell(row=wt_start, column=1, value="Tier\\Row")
        for r in range(n_rows_bp):
            cell = ws_bp.cell(row=wt_start, column=r+2, value=f"R{r}")
            cell.font = header_font
            cell.fill = header_fill_ppo
            cell.alignment = center_align
        for t_d in range(n_tiers_bp):
            t_a = n_tiers_bp - 1 - t_d
            dr = wt_start + 1 + t_d
            ws_bp.cell(row=dr, column=1, value=f"T{t_a}").font = Font(bold=True)
            for r in range(n_rows_bp):
                cell = ws_bp.cell(row=dr, column=r+2)
                cell.alignment = center_align
                cell.border = thin_border
                h = int(stack_h[r])
                if t_a < h:
                    wv = float(wt_grid[r, t_a])
                    cell.value = round(wv, 1)
                    ratio = (wv - MIN_WT) / (MAX_WT - MIN_WT + 1e-9)
                    if ratio > 0.7:
                        cell.fill = PatternFill("solid", fgColor="FF6B6B")
                    elif ratio > 0.4:
                        cell.fill = PatternFill("solid", fgColor="FFD93D")
                    else:
                        cell.fill = PatternFill("solid", fgColor="C1FFC1")
                else:
                    cell.value = "—"
                    cell.fill = pod_fills[0]

        # 선적 순서 리스트
        ls = wt_start + 1 + n_tiers_bp + 2
        ws_bp.cell(row=ls, column=1, value="【선적 순서】").font = Font(bold=True, size=11)
        ls += 1
        for ci, h in enumerate(["순번", "POD", "POD Name", "Weight(MT)"], 1):
            cell = ws_bp.cell(row=ls, column=ci, value=h)
            cell.font = header_font
            cell.fill = header_fill_gen
            cell.alignment = center_align
        for i in range(len(bp["ctns_wt"])):
            dr = ls + 1 + i
            ws_bp.cell(row=dr, column=1, value=i+1).alignment = center_align
            ws_bp.cell(row=dr, column=2, value=int(bp["ctns_pod"][i])).alignment = center_align
            ws_bp.cell(row=dr, column=3, value=POD_NAMES.get(int(bp["ctns_pod"][i]),"?")).alignment = center_align
            ws_bp.cell(row=dr, column=4, value=round(float(bp["ctns_wt"][i]),1)).alignment = center_align

        ws_bp.column_dimensions["A"].width = 12
        for r in range(n_rows_bp):
            ws_bp.column_dimensions[get_column_letter(r+2)].width = 12

    # ══════ Config Sheet ══════
    ws_cfg = wb.create_sheet("Config")
    ws_cfg.cell(row=1, column=1, value="Parameter").font = Font(bold=True)
    ws_cfg.cell(row=1, column=2, value="Value").font = Font(bold=True)
    items = [
        ("Version", "v6 (Same-Tier Band reward)"),
        ("Seed", seed), ("N_POD", N_POD), ("MAX_ROWS", MAX_ROWS),
        ("OBS_DIM", OBS_DIM),
        ("Container Weight", f"{MIN_WT:.0f}–{MAX_WT:.0f} MT"),
        ("Max Column Weight", f"{MAX_COL_WT:.0f} MT"),
        ("Fill Ratio", "100%"),
        ("Steps/Round", config["timesteps_per_round"]),
        ("Total Rounds", config["total_rounds"]),
        ("Eval Episodes", config["eval_episodes"]),
        ("Loading Order", "POD desc → Weight desc"),
        ("Algorithm", "PPO (only)"),
        ("R13 Definition", "Same-Tier Band bonus (replaces center row bonus)"),
    ]
    for i, (k, v) in enumerate(items, 2):
        ws_cfg.cell(row=i, column=1, value=k)
        ws_cfg.cell(row=i, column=2, value=str(v))

    r = len(items) + 3
    ws_cfg.cell(row=r, column=1, value="Reward Weights").font = Font(bold=True)
    r += 1
    for rw_name, rw_val in config["rw"].items():
        ws_cfg.cell(row=r, column=1, value=rw_name)
        ws_cfg.cell(row=r, column=2, value=rw_val)
        r += 1

    xlsx_path = f"{logger.log_dir}/ppo_results_v6.xlsx"
    wb.save(xlsx_path)
    print(f"\n  📊 Excel saved: {xlsx_path}")
    return xlsx_path


# ─────────────────────────────────────────────────────────────────────
# ★ v8 신규: 랜덤 입력 데이터 (eval 컨테이너 시퀀스) Excel 저장
# ─────────────────────────────────────────────────────────────────────
def save_random_input_data_to_excel(
    ppo_res: Dict, config: dict,
    seed: int, logger: ExperimentLogger,
) -> str:
    """학습/평가에 사용된 랜덤 컨테이너 데이터를 Excel로 저장.

    각 라운드별 시트에 30개 평가 에피소드의 컨테이너 시퀀스를 기록.
    컨테이너 생성은 deterministic (seed 고정) 이므로 사후 재생성 가능.
      - eval_seed = seed + ep * 17  (env.reset 시 사용된 것과 동일)
      - 정렬: POD desc → Weight desc (env 내부 lexsort 와 동일)

    추가로 'Overview' 와 'POD_Distribution' 통계 시트 포함.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", fgColor="404040")
    accent_fill = PatternFill("solid", fgColor="2166AC")
    center      = Alignment(horizontal="center", vertical="center")
    thin        = Border(left=Side(style="thin"),  right=Side(style="thin"),
                         top=Side(style="thin"),   bottom=Side(style="thin"))

    n_eval = config["eval_episodes"]

    # ══════ Sheet 1: Overview ══════
    ws = wb.active
    ws.title = "Overview"
    title = ws.cell(row=1, column=1,
        value=f"Random Input Data — PPO v8 | seed={seed} | "
              f"{config['total_rounds']} rounds × {n_eval} eval episodes")
    title.font = Font(bold=True, size=12)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)

    headers_ov = ["Round", "Level", "Grid", "N_Containers",
                  "N_Eval_Episodes", "Eval_Seed_Range", "Avg_Weight (MT)"]
    for col, h in enumerate(headers_ov, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = header_font; c.fill = header_fill; c.alignment = center

    for rnd_idx in range(config["total_rounds"]):
        lv     = ppo_res["levels"][rnd_idx]
        lv_cfg = config["levels"][lv]
        n_cont = lv_cfg["n_containers"]
        # 평균 무게 (모든 eval 에피소드의 컨테이너 무게 평균)
        all_weights = []
        for ep in range(n_eval):
            ep_seed = seed + ep * 17
            rng = np.random.default_rng(ep_seed)
            wts = rng.uniform(MIN_WT, MAX_WT, n_cont).astype(np.float32)
            all_weights.extend(wts.tolist())
        avg_w = float(np.mean(all_weights))

        seed_range = f"{seed} ~ {seed + (n_eval - 1) * 17} (step 17)"
        row_n = 4 + rnd_idx
        ws.cell(row=row_n, column=1, value=rnd_idx + 1).alignment = center
        ws.cell(row=row_n, column=2, value=f"Lv{lv}").alignment = center
        ws.cell(row=row_n, column=3,
                value=f"{lv_cfg['n_rows']}×{lv_cfg['n_tiers']}").alignment = center
        ws.cell(row=row_n, column=4, value=n_cont).alignment = center
        ws.cell(row=row_n, column=5, value=n_eval).alignment = center
        ws.cell(row=row_n, column=6, value=seed_range).alignment = center
        ws.cell(row=row_n, column=7, value=round(avg_w, 2)).alignment = center

    # 컬럼 너비
    for col_letter, w in [("A", 8), ("B", 8), ("C", 10), ("D", 14),
                           ("E", 16), ("F", 28), ("G", 16)]:
        ws.column_dimensions[col_letter].width = w

    # 안내문
    ws.cell(row=4 + config["total_rounds"] + 2, column=1,
        value="• 정렬 규칙: POD desc → Weight desc (env._reset() 내부 lexsort)").font = Font(italic=True)
    ws.cell(row=4 + config["total_rounds"] + 3, column=1,
        value=f"• POD 범위: 1 ~ {N_POD}  |  Weight 범위: {MIN_WT:.0f} ~ {MAX_WT:.0f} MT").font = Font(italic=True)
    ws.cell(row=4 + config["total_rounds"] + 4, column=1,
        value="• Eval seed 공식: eval_seed = base_seed + ep_idx × 17 (deterministic)").font = Font(italic=True)

    # ══════ POD 색상 (results sheet 와 동일) ══════
    pod_fills = {
        1: PatternFill("solid", fgColor="4393C3"),
        2: PatternFill("solid", fgColor="F4A582"),
        3: PatternFill("solid", fgColor="D6604D"),
        4: PatternFill("solid", fgColor="92C5DE"),
        5: PatternFill("solid", fgColor="2166AC"),
        6: PatternFill("solid", fgColor="B2182B"),
    }
    white_font = Font(bold=False, color="FFFFFF", size=9)
    black_font = Font(bold=False, color="000000", size=9)

    # ══════ Sheet 2~ : 라운드별 입력 데이터 ══════
    for rnd_idx in range(config["total_rounds"]):
        lv     = ppo_res["levels"][rnd_idx]
        n_cont = config["levels"][lv]["n_containers"]
        ws_r   = wb.create_sheet(f"R{rnd_idx+1}_Lv{lv}_Inputs")

        # 제목
        ws_r.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
        title_c = ws_r.cell(row=1, column=1,
            value=f"Round {rnd_idx+1}  Lv{lv}  ({n_cont} containers × "
                  f"{n_eval} eval episodes = {n_cont * n_eval:,} rows)")
        title_c.font = Font(bold=True, size=12)

        # 헤더
        headers_r = ["Eval_Episode", "Eval_Seed", "Loading_Order",
                     "POD", "POD_Name", "Weight_MT"]
        for col, h in enumerate(headers_r, 1):
            c = ws_r.cell(row=3, column=col, value=h)
            c.font = header_font; c.fill = header_fill; c.alignment = center

        # 데이터: 각 eval 에피소드의 컨테이너 시퀀스 재생성
        row_n = 4
        for ep in range(n_eval):
            ep_seed = seed + ep * 17
            # env.reset() 의 RNG 로직과 동일하게 재현
            rng = np.random.default_rng(ep_seed)
            raw_wt  = rng.uniform(MIN_WT, MAX_WT, n_cont).astype(np.float32)
            raw_pod = rng.integers(1, N_POD + 1, n_cont).astype(np.int32)
            sort_keys = np.lexsort((-raw_wt, -raw_pod))
            ctns_wt   = raw_wt[sort_keys]
            ctns_pod  = raw_pod[sort_keys]

            for i in range(n_cont):
                ws_r.cell(row=row_n, column=1, value=ep + 1).alignment = center
                ws_r.cell(row=row_n, column=2, value=ep_seed).alignment = center
                ws_r.cell(row=row_n, column=3, value=i + 1).alignment = center
                p = int(ctns_pod[i])
                pcell = ws_r.cell(row=row_n, column=4, value=p)
                pcell.alignment = center
                pcell.fill = pod_fills.get(p, PatternFill())
                pcell.font = white_font if p >= 3 else black_font
                ws_r.cell(row=row_n, column=5,
                          value=POD_NAMES.get(p, "?")).alignment = center
                ws_r.cell(row=row_n, column=6,
                          value=round(float(ctns_wt[i]), 2)).alignment = center
                row_n += 1

        # 컬럼 너비
        for col_letter, w in [("A", 14), ("B", 14), ("C", 14),
                              ("D", 8), ("E", 14), ("F", 12)]:
            ws_r.column_dimensions[col_letter].width = w

        # ── 라운드 통계 (POD 분포) ──
        stat_start = row_n + 2
        ws_r.cell(row=stat_start, column=1,
                  value="【POD Distribution (이 라운드 전체)】").font = Font(bold=True, size=11)
        # POD 카운트 집계
        pod_counts = {p: 0 for p in range(1, N_POD + 1)}
        for ep in range(n_eval):
            ep_seed = seed + ep * 17
            rng = np.random.default_rng(ep_seed)
            raw_pod = rng.integers(1, N_POD + 1, n_cont)
            for p in raw_pod:
                pod_counts[int(p)] += 1
        total = sum(pod_counts.values())

        ws_r.cell(row=stat_start + 1, column=1, value="POD").font = Font(bold=True)
        ws_r.cell(row=stat_start + 1, column=2, value="POD Name").font = Font(bold=True)
        ws_r.cell(row=stat_start + 1, column=3, value="Count").font = Font(bold=True)
        ws_r.cell(row=stat_start + 1, column=4, value="Ratio (%)").font = Font(bold=True)
        for i, (p, cnt) in enumerate(sorted(pod_counts.items())):
            r = stat_start + 2 + i
            pc = ws_r.cell(row=r, column=1, value=p)
            pc.alignment = center
            pc.fill = pod_fills.get(p, PatternFill())
            pc.font = white_font if p >= 3 else black_font
            ws_r.cell(row=r, column=2,
                      value=POD_NAMES.get(p, "?")).alignment = center
            ws_r.cell(row=r, column=3, value=cnt).alignment = center
            ws_r.cell(row=r, column=4,
                      value=round(cnt / total * 100, 1)).alignment = center

    # ══════ 저장 ══════
    inputs_path = f"{logger.log_dir}/ppo_random_inputs_v8.xlsx"
    wb.save(inputs_path)
    print(f"  📊 Random Inputs Excel saved: {inputs_path}")
    return inputs_path


# ── 보상 항목별 분석 — PPO ─────────────────────────────────────────
def _evaluate_reward_components(
    model, level: int, config: dict, seed: int,
    n_eval: int = 30,
) -> Dict[str, List[float]]:
    """보상 항목별 누적값을 에피소드별로 수집."""
    comp_data: Dict[str, List[float]] = defaultdict(list)
    total_rewards: List[float] = []

    for ep in range(n_eval):
        env = make_env(level=level, config=config)
        obs, _ = env.reset(seed=seed + ep * 17)
        done = False
        ep_reward = 0.0
        while not done:
            act, _ = model.predict(obs, deterministic=True)
            obs, r, term, trunc, _ = env.step(act)
            done = term or trunc
            ep_reward += r

        for key, val in env.metrics.reward_components.items():
            comp_data[key].append(val)
        total_rewards.append(ep_reward)

    comp_data["Total"] = total_rewards
    return dict(comp_data)


def print_reward_component_table(
    ppo_res: Dict, config: dict, seed: int,
) -> Dict:
    """최종 라운드의 PPO 보상 항목별 테이블 출력 (v6: R13 의미 변경)."""
    n_eval = config["eval_episodes"]

    print("\n  📊 Evaluating reward components...")
    ppo_comp = _evaluate_reward_components(
        ppo_res["trained_model"], ppo_res["levels"][-1],
        config, seed, n_eval)

    comp_desc = {
        "R1_valid"          : "R1: Valid placement       (+)",
        "R2_stack_full"     : "R2: Stack-full penalty    (−)",
        "R3_overstow"       : "R3: Overstow penalty      (−)",
        "R4_order"          : "R4: Discharge order        (+)",
        "R5_weight_bal"     : "R5: Weight balance         (+)",
        "R6_cog"            : "R6: COG deviation          (−)",
        "R7_completion"     : "R7: Completion bonus       (+)",
        "R8_pod_band"       : "R8: POD band quality       (+)",
        "R9_col_wt"         : "R9: Col weight penalty     (−)",
        "R10_tier_match"    : "R10: Tier POD match        (+)",
        "R11_wt_inversion"  : "R11: Wt inversion          (−)",
        "R12_col_order"     : "R12: Col order perfection  (+)",
        "R13_tier_band"     : "R13: Same-Tier Band       (+/−)",   # ★ v6
        "R14_empty_row"     : "R14: Empty row penalty     (−)",
        "Total"             : "Total Episode Reward",
    }
    comp_order = list(comp_desc.keys())

    print("\n" + "═" * 70)
    print("  TABLE: Reward Component Breakdown — PPO (Final Round, v6)")
    print("═" * 70)
    print(f"  {'Component':<35} {'Mean':>12} {'Std':>10}")
    print("  " + "─" * 66)

    for key in comp_order:
        pv = np.array(ppo_comp.get(key, [0]))
        desc = comp_desc.get(key, key)
        if key == "Total":
            print("  " + "─" * 66)
        print(f"  {desc:<35} {np.mean(pv):>+12.2f} {np.std(pv):>10.2f}")

    print("═" * 70)
    return ppo_comp


# ── 보상 항목별 바 차트 — PPO (v6: R13 라벨 변경) ──────────────────
def plot_reward_components(
    ppo_comp: Dict,
    logger: ExperimentLogger,
) -> None:
    """Fig.6: PPO 보상 항목별 바 차트."""
    _scie_style()

    comp_order = [
        "R1_valid", "R2_stack_full", "R3_overstow", "R4_order",
        "R5_weight_bal", "R6_cog", "R7_completion", "R8_pod_band", "R9_col_wt",
        "R10_tier_match", "R11_wt_inversion", "R12_col_order",
        "R13_tier_band", "R14_empty_row",                            # ★ v6
    ]
    short_names = ["R1\nValid", "R2\nStack\nFull", "R3\nOver-\nstow",
                   "R4\nOrder", "R5\nWt\nBal.", "R6\nCOG",
                   "R7\nCompl.", "R8\nPOD\nBand", "R9\nCol\nWT",
                   "R10\nTier\nMatch", "R11\nWt\nInv.", "R12\nCol\nOrd",
                   "R13\nTier\nBand", "R14\nEmpty\nRow"]              # ★ v6 라벨 변경

    x = np.arange(len(comp_order))
    w = 0.55

    fig, ax = plt.subplots(figsize=(16, 6))
    fig.suptitle(
        "Fig. 6  PPO — Reward Component Breakdown (Final Round, v6)",
        fontsize=11, fontweight="bold")

    means = [float(np.mean(ppo_comp.get(k, [0]))) for k in comp_order]
    stds  = [float(np.std(ppo_comp.get(k, [0])))  for k in comp_order]
    bars = ax.bar(x, means, w, yerr=stds, label="PPO",
                  color=CLR["PPO"], alpha=0.85, edgecolor="white",
                  capsize=2, error_kw={"lw": 0.8})
    for bar in bars:
        h = bar.get_height()
        if abs(h) > 0.5:
            va = "bottom" if h >= 0 else "top"
            ax.text(bar.get_x() + bar.get_width()/2., h + (0.5 if h >= 0 else -0.5),
                    f"{h:.1f}", ha="center", va=va, fontsize=7)

    ax.set_xlabel("Reward Component")
    ax.set_ylabel("Cumulative Reward per Episode")
    ax.set_xticks(x)
    ax.set_xticklabels(short_names, fontsize=7)
    ax.axhline(0, color="gray", ls="--", lw=0.8, alpha=0.5)
    ax.legend()
    ax.grid(axis="y", alpha=0.3)

    plt.tight_layout()
    logger.save_fig(fig, "fig6_reward_components.png")
    plt.show(); plt.close(fig)


print("✅ [F] All visualization & export functions defined (PPO only, v6)")
print()
print("  All sections loaded. Proceed to Cell 3 (CONFIG) → Cell 4 (Experiment).")


# ===== CONFIG / POLICIES =====
# ══════════════════════════════════════════════════════════════════
#  CONFIG v15: 보상 계층(필수/안정/효율) 4-정책 비교 (ES / SF / EF / BL)
#   · ★★ v15 변경:
#       ① 보상 항목 계층 재분류
#          - 필수: R1, R2, R3, R7, R9, R11  (모든 정책 항상 활성)
#          - 안정: R5, R6
#          - 효율: R4, R8, R10, R12, R13, R14, R15
#       ② 조정 가중치 적용 (안정 ×2 상향, 효율 ×2/3 하향)
#          - R5: 3.0→6.0, R6: −4.0→−8.0
#          - R4: 3.0→2.0, R8: 2.0→1.33, R10: 4.0→2.67, R12: 3.0→2.0,
#            R13: 2.0→1.33 (−1.5→−1.0), R14: −3.0→−2.0, R15: −4.0→−2.67
#       ③ 4-정책 커리큘럼 (각 정책 전 라운드 고정 보상 집합)
#          ① ES: 필수만 | ② SF: 필수+안정 | ③ EF: 필수+효율 | ④ BL: 전체
#   · 총 학습량: 4 정책 × 125K eps = 500K eps ≈ 32.3M steps
#   · 학습 종료 기준: episodes_per_level[lv] 에피소드 도달 시 callback이 종료
# ══════════════════════════════════════════════════════════════════
CONFIG = {
    "experiment_name"    : "single_bay_6pod_ppo_v15_4policy",
    "seed"               : GLOBAL_SEED,

    # ── 실험 제어 ────────────────────────────────────────────────
    "total_rounds"       : 4,
    "eval_episodes"      : 30,
    "level_up_threshold" : 0.0,

    # ★★★ v8 신규: 레벨별 "에피소드" 예산 (스텝이 아닌 에피소드로 종료) ★★★
    # 학습은 정확히 episodes_per_level[lv] 에피소드를 채우면 종료된다.
    # SB3 model.learn 의 total_timesteps 는 안전 상한으로만 사용되며,
    # 실제 종료는 EpisodeMetricsCallback 이 episode 카운트를 보고 결정한다.
    "episodes_per_level": {
        1: 20_000,    # Lv1:  16 cont →   320K steps   (4.0%)
        2: 25_000,    # Lv2:  36 cont →   900K steps  (11.2%)
        3: 32_000,    # Lv3:  64 cont → 2,048K steps  (25.4%)
        4: 48_000,    # Lv4: 100 cont → 4,800K steps  (59.5%)  ★ 주 실험, 2,343 PPO updates
    },
    # Total: 125,000 episodes / policy  →  ~8,068,000 steps / policy
    "max_steps_safety_factor": 2.0,    # 안전 상한 = episodes × n_cont × factor
    "timesteps_per_round": 200_000,    # legacy (현재 사용 안 함, 호환성 유지)

    # ── 커리큘럼 레벨 (4단계, 100% 적재율) ─────────────────────
    "levels": {
        1: {"n_containers": 16, "n_rows": 4, "n_tiers": 4,
            "desc": "Lv1: Small  (16 GP-40 | 4×4, 100%)"},
        2: {"n_containers": 36, "n_rows": 6, "n_tiers": 6,
            "desc": "Lv2: Medium (36 GP-40 | 6×6, 100%)"},
        3: {"n_containers": 64, "n_rows": 8, "n_tiers": 8,
            "desc": "Lv3: Large  (64 GP-40 | 8×8, 100%)"},
        4: {"n_containers": 100, "n_rows": 10, "n_tiers": 10,
            "desc": "Lv4: XLarge (100 GP-40 | 10×10, 100%)"},
    },

    # ── 보상 가중치 (R1~R15, ★ v15: 조정 가중치 — 계층: 필수/안정/효율) ──
    "rw": {
        "valid"               :    1.0,   # R1  [필수] 유지
        "stack_full"          :  -10.0,   # R2  [필수] 유지
        "overstow"            :  -15.0,   # R3  [필수] 유지
        "order_bonus"         :    2.0,   # R4  [효율] ★ v15: 3.0 → 2.0
        "weight_balance"      :    6.0,   # R5  [안정] ★ v15: 3.0 → 6.0
        "cog_penalty"         :   -8.0,   # R6  [안정] ★ v15: -4.0 → -8.0
        "completion"          :   80.0,   # R7  [필수] 유지 (×VPR)
        "pod_band_quality"    :    1.33,  # R8  [효율] ★ v15: 2.0 → 1.33
        "col_wt_penalty"      :   -5.0,   # R9  [필수] 유지
        "tier_pod_match"      :    2.67,  # R10 [효율] ★ v15: 4.0 → 2.67
        "wt_inversion_penalty":   -6.0,   # R11 [필수] 유지
        "col_order_bonus"     :    2.0,   # R12 [효율] ★ v15: 3.0 → 2.0
        "tier_band_bonus"     :    1.33,  # R13 [효율] ★ v15: 2.0 → 1.33
        "tier_band_penalty"   :   -1.0,   # R13b[효율] ★ v15: -1.5 → -1.0 (R13 비례 축소)
        "empty_row_penalty"   :   -2.0,   # R14 [효율] ★ v15: -3.0 → -2.0 (×빈 Row 수)
        "vstack_pod_penalty"  :   -2.67,  # R15 [효율] ★ v15: -4.0 → -2.67
    },

    # ── PPO 하이퍼파라미터 ───────────────────────────────────────
    "ppo": {
        "lr"        : 3e-4,
        "gamma"     : 0.99,
        "n_steps"   : 2048,
        "batch_size": 64,
        "n_epochs"  : 10,
        "clip_range": 0.2,
        "ent_coef"  : 0.01,
        "gae_lambda": 0.95,
        "net_arch"  : [512, 256, 128],
    },
}

# ── 설정 확인 출력 ──
print("✅ CONFIG loaded (PPO only, v15: 4-policy ES/SF/EF/BL + 계층 가중치 + 8M steps/policy)")
print()
print(f"  Experiment : {CONFIG['experiment_name']}")
print(f"  Seed       : {GLOBAL_SEED} (fixed)")
print(f"  Rounds     : {CONFIG['total_rounds']}")
print()
print(f"  ★ v8: 레벨별 에피소드 예산 (스텝이 아닌 에피소드로 학습 종료)")
print(f"  {'Level':<6} {'n_cont':>7} {'Episodes':>10} {'≈MaxSteps':>12}")
print("  " + "─" * 44)
total_eps    = 0
total_max_ts = 0
sf = CONFIG.get("max_steps_safety_factor", 2.0)
for lv, eps in CONFIG["episodes_per_level"].items():
    n_cont = CONFIG["levels"][lv]["n_containers"]
    max_ts = int(eps * n_cont * sf)
    total_eps    += eps
    total_max_ts += max_ts
    print(f"  Lv{lv}    {n_cont:>7}    {eps:>9,}    {max_ts:>11,}")
print("  " + "─" * 44)
print(f"  {'Total':<14} {total_eps:>9,}  ≤ {total_max_ts:>11,}")
print(f"  Safety factor: ×{sf}  (실제 학습은 callback이 목표 episode 도달 시 종료)")
print()
print(f"  보상 구조: v15 — 계층 분류(필수/안정/효율) + 조정 가중치 (안정 ×2, 효율 ×2/3)")
print(f"  Container : {MIN_WT:.0f}–{MAX_WT:.0f} MT  |  Max Col WT: {MAX_COL_WT:.0f} MT")
print()
print("  Curriculum levels (100% fill):")
for lv, lvc in CONFIG["levels"].items():
    cap = lvc["n_rows"] * lvc["n_tiers"]
    util = lvc["n_containers"] / cap * 100
    print(f"    Lv{lv}: {lvc['desc']} ({util:.0f}%)")
print()

# ── 환경 검증 ──
_env_test = SingleBayStowageEnv(curriculum_level=1, config=CONFIG)
_obs, _   = _env_test.reset(seed=42)
assert _obs.shape == (OBS_DIM,), f"Obs dim mismatch: {_obs.shape} != {OBS_DIM}"
_env_test.step(_env_test.action_space.sample())
_pods = _env_test.ctns_pod
print(f"✅ PPO Environment check passed (obs_dim={OBS_DIM})")
print(f"  Loading order: {list(_pods[:8])}... (POD desc ✅)")

# ── R13 (Same-Tier Band) 동작 검증 ──
_env_test2 = SingleBayStowageEnv(curriculum_level=1, config=CONFIG)
_env_test2.reset(seed=42)
first_pod = int(_env_test2.ctns_pod[0])
target_t  = _env_test2._get_target_tier_for_pod(first_pod)
assert target_t == 0, f"First container target tier should be 0 (got {target_t})"
print(f"✅ R13 Same-Tier Band check: first POD={first_pod}, target_tier={target_t} ✓")

# ── 레벨별 예산 검증 (v8: episode 기반) ──
assert "episodes_per_level" in CONFIG, "episodes_per_level missing"
assert all(lv in CONFIG["episodes_per_level"] for lv in CONFIG["levels"]), \
       "All curriculum levels must have an episodes_per_level entry"
print(f"✅ Per-level EPISODE budget configured for all "
      f"{len(CONFIG['levels'])} levels (callback-based termination)")

# ── ★ v15: 조정 가중치 변경 확인 (v14 대비) ──
print()
print("  ── ★ v15 조정 가중치 (v14 → v15) ──")
print(f"  {'ID':<6} {'Key':<22} {'v14':>8} {'v15':>8} {'배율':>6}")
print("  " + "─" * 56)
_v14_w = {
    "order_bonus": 3.0, "weight_balance": 3.0, "cog_penalty": -4.0,
    "pod_band_quality": 2.0, "tier_pod_match": 4.0, "col_order_bonus": 3.0,
    "tier_band_bonus": 2.0, "tier_band_penalty": -1.5,
    "empty_row_penalty": -3.0, "vstack_pod_penalty": -4.0,
}
_rid = {"order_bonus": "R4", "weight_balance": "R5", "cog_penalty": "R6",
        "pod_band_quality": "R8", "tier_pod_match": "R10",
        "col_order_bonus": "R12", "tier_band_bonus": "R13a",
        "tier_band_penalty": "R13b", "empty_row_penalty": "R14",
        "vstack_pod_penalty": "R15"}
for k, v14 in _v14_w.items():
    v15 = CONFIG["rw"][k]
    ratio = abs(v15 / v14) if v14 != 0 else float('inf')
    print(f"  {_rid[k]:<6} {k:<22} {v14:>+7.2f} {v15:>+7.2f} {ratio:>5.2f}×")
del _env_test, _env_test2, _obs



# ══════════════════════════════════════════════════════════════════
#  ★ v15 NEW: 보상 계층(필수/안정/효율) 기반 4-정책 정의
#   · 필수(Essential)  : R1, R2, R3, R7, R9, R11 — 모든 정책 항상 활성
#   · 안정(Stability)  : R5, R6
#   · 효율(Efficiency) : R4, R8, R10, R12, R13(a/b), R14, R15
# ══════════════════════════════════════════════════════════════════

# ── 보상 그룹 분류 (★ v15) ──
ESSENTIAL_KEYS = [
    "valid",                # R1  유효 적재 보너스
    "stack_full",           # R2  스택 가득참 패널티
    "overstow",             # R3  과적(overstow) 패널티
    "completion",           # R7  에피소드 완료 보너스
    "col_wt_penalty",       # R9  열 중량제한 초과 패널티
    "wt_inversion_penalty", # R11 무게 역전 패널티
]
STABILITY_KEYS = [
    "weight_balance",       # R5  중량 균형 보너스
    "cog_penalty",          # R6  COG 편차 패널티
]
EFFICIENCY_KEYS = [
    "order_bonus",          # R4  올바른 하역순서 보너스
    "pod_band_quality",     # R8  POD 수평밴드 품질 보너스
    "tier_pod_match",       # R10 같은 층 같은 POD 보너스
    "col_order_bonus",      # R12 열별 하역순서 완벽도
    "tier_band_bonus",      # R13a Same-Tier Band 보너스
    "tier_band_penalty",    # R13b Same-Tier Band 패널티
    "empty_row_penalty",    # R14 빈 Row 패널티
    "vstack_pod_penalty",   # R15 수직 동일-POD 인접 패널티
]

# 분류 무결성 검증: 세 그룹의 합집합 = CONFIG["rw"] 전체 키
_all_keys = set(ESSENTIAL_KEYS) | set(STABILITY_KEYS) | set(EFFICIENCY_KEYS)
assert _all_keys == set(CONFIG["rw"].keys()), \
    f"보상 그룹 분류 누락/중복: {_all_keys ^ set(CONFIG['rw'].keys())}"
assert not (set(ESSENTIAL_KEYS) & set(STABILITY_KEYS)) \
   and not (set(ESSENTIAL_KEYS) & set(EFFICIENCY_KEYS)) \
   and not (set(STABILITY_KEYS) & set(EFFICIENCY_KEYS)), "그룹 간 중복 키 존재"


def make_phase_weights(base_w: dict, mode: str) -> dict:
    """학습 모드별 보상 가중치 생성 (★ v15: 계층 기반 4모드).

    Args:
        base_w: BASE 보상 가중치 (CONFIG["rw"])
        mode  :
          'ESSENTIAL_ONLY'       — ① 필수 항목만 활성
          'ESSENTIAL_STABILITY'  — ② 필수 + 안정×0.7 + 효율×0.3 (★ 7:3 비율)
          'ESSENTIAL_EFFICIENCY' — ③ 필수 + 안정×0.3 + 효율×0.7 (★ 3:7 비율)
          'BASE'                 — ④ 필수 + 안정 + 효율 전체 활성
          (legacy alias 호환: 'STABILITY_ONLY'→②, 'EFFICIENCY_ONLY'→③)

    Returns:
        modified weight dict (same keys, 비활성/축소 그룹은 비율 가중치 적용)
    """
    # legacy alias 호환 (v11~v14 코드와의 하위 호환)
    _alias = {"STABILITY_ONLY":  "ESSENTIAL_STABILITY",
              "EFFICIENCY_ONLY": "ESSENTIAL_EFFICIENCY"}
    mode = _alias.get(mode, mode)

    w = dict(base_w)
    if mode == "ESSENTIAL_ONLY":
        for k in STABILITY_KEYS + EFFICIENCY_KEYS:
            w[k] = 0.0
    elif mode == "ESSENTIAL_STABILITY":
        # ★ 7:3 — SF: 필수(유지) + 안정×0.7 + 효율×0.3
        for k in STABILITY_KEYS:
            w[k] = base_w[k] * 0.7
        for k in EFFICIENCY_KEYS:
            w[k] = base_w[k] * 0.3
    elif mode == "ESSENTIAL_EFFICIENCY":
        # ★ 3:7 — EF: 필수(유지) + 안정×0.3 + 효율×0.7
        for k in STABILITY_KEYS:
            w[k] = base_w[k] * 0.3
        for k in EFFICIENCY_KEYS:
            w[k] = base_w[k] * 0.7
    elif mode != "BASE":
        raise ValueError(f"Unknown mode: {mode!r}")
    return w


# ── 정책 정의 (★ v15: 4-way — 보상 그룹별 ablation) ──
#   각 정책은 4 라운드(커리큘럼 Lv1→Lv4) 내내 *고정된* 보상 집합으로 학습한다.
#   (phase1_mode == phase2_mode 로 설정 → 기존 round 기반 phase 메커니즘 재사용)
#   ① ES: 필수만            ② SF: 필수+안정
#   ③ EF: 필수+효율         ④ BL: 필수+안정+효율 (전체)
POLICIES = {
    "ES": {
        "desc":          "Essential-Only (필수 R1,R2,R3,R7,R9,R11 only, all rounds)",
        "kor":           "① 필수 항목만 학습",
        "color":         "#2CA02C",   # green
        "marker":        "D",
        "phase1_rounds": [1, 2],
        "phase2_rounds": [3, 4],
        "phase1_mode":   "ESSENTIAL_ONLY",
        "phase2_mode":   "ESSENTIAL_ONLY",   # ★ 전 라운드 동일 — phase 전환 없음
        "is_baseline":   True,               # 시각화에서 P1→P2 boundary 미표시
    },
    "SF": {
        "desc":          "Safety-First (필수 + 안정×0.7 + 효율×0.3, all rounds)",
        "kor":           "② 필수 + 안정×0.7 + 효율×0.3 (7:3)",
        "color":         "#2E75B6",   # blue
        "marker":        "o",
        "phase1_rounds": [1, 2],
        "phase2_rounds": [3, 4],
        "phase1_mode":   "ESSENTIAL_STABILITY",
        "phase2_mode":   "ESSENTIAL_STABILITY",
        "is_baseline":   True,
    },
    "EF": {
        "desc":          "Efficiency-First (필수 + 안정×0.3 + 효율×0.7, all rounds)",
        "kor":           "③ 필수 + 안정×0.3 + 효율×0.7 (3:7)",
        "color":         "#C00000",   # red
        "marker":        "s",
        "phase1_rounds": [1, 2],
        "phase2_rounds": [3, 4],
        "phase1_mode":   "ESSENTIAL_EFFICIENCY",
        "phase2_mode":   "ESSENTIAL_EFFICIENCY",
        "is_baseline":   True,
    },
    "BL": {
        "desc":          "Baseline/Full (필수 + 안정 + 효율 — all R1~R15, all rounds)",
        "kor":           "④ 필수 + 안정 + 효율 전체 학습",
        "color":         "#7F7F7F",   # gray
        "marker":        "^",
        "phase1_rounds": [1, 2],
        "phase2_rounds": [3, 4],
        "phase1_mode":   "BASE",
        "phase2_mode":   "BASE",
        "is_baseline":   True,
    },
}

# R 항목 매핑 (출력용)
R_MAP = {
    "valid": "R1", "stack_full": "R2", "overstow": "R3", "order_bonus": "R4",
    "weight_balance": "R5", "cog_penalty": "R6", "completion": "R7",
    "pod_band_quality": "R8", "col_wt_penalty": "R9", "tier_pod_match": "R10",
    "wt_inversion_penalty": "R11", "col_order_bonus": "R12",
    "tier_band_bonus": "R13a", "tier_band_penalty": "R13b",
    "empty_row_penalty": "R14", "vstack_pod_penalty": "R15",
}

# 계층 매핑 (출력용)
TIER_MAP = {}
for _k in ESSENTIAL_KEYS:  TIER_MAP[_k] = "필수"
for _k in STABILITY_KEYS:  TIER_MAP[_k] = "안정"
for _k in EFFICIENCY_KEYS: TIER_MAP[_k] = "효율"

# ── 정책 검증 출력 ──
_n_rw = len(CONFIG["rw"])
print()
print("  ── ★ v15 보상 계층 4-정책 (전 라운드 고정 보상 집합) ──")
print(f"  {'정책':<6} {'활성 보상 집합 (R1~R4 전 라운드)':<26} {'활성수':>8}   {'설명'}")
print("  " + "─" * 90)
for pn, ps in POLICIES.items():
    p_w = make_phase_weights(CONFIG["rw"], ps["phase1_mode"])
    p_active = sum(1 for v in p_w.values() if v != 0)
    print(f"  {pn:<6} {ps['phase1_mode']:<26} ({p_active:2}/{_n_rw})   {ps['kor']}")

print()
print("  ── 정책별 활성 보상 가중치 비교 (0.00 = 비활성) ──")
print(f"  {'R':<6} {'계층':<5} {'Key':<22} {'ES':>8} {'SF':>8} {'EF':>8} {'BL':>8}")
print("  " + "─" * 76)
_es_w = make_phase_weights(CONFIG["rw"], "ESSENTIAL_ONLY")
_sf_w = make_phase_weights(CONFIG["rw"], "ESSENTIAL_STABILITY")
_ef_w = make_phase_weights(CONFIG["rw"], "ESSENTIAL_EFFICIENCY")
_bl_w = make_phase_weights(CONFIG["rw"], "BASE")
for k in CONFIG["rw"]:
    print(f"  {R_MAP.get(k, '?'):<6} {TIER_MAP.get(k, '?'):<5} {k:<22} "
          f"{_es_w[k]:>+8.2f} {_sf_w[k]:>+8.2f} {_ef_w[k]:>+8.2f} {_bl_w[k]:>+8.2f}")
print("  " + "─" * 76)
print(f"  활성수: ES={sum(1 for v in _es_w.values() if v != 0)}, "
      f"SF={sum(1 for v in _sf_w.values() if v != 0)}, "
      f"EF={sum(1 for v in _ef_w.values() if v != 0)}, "
      f"BL={sum(1 for v in _bl_w.values() if v != 0)}  (/{_n_rw} keys, R13은 a/b 2키)")


# ===== CBData (TRAIN 저장 / AGGREGATE 로드 공용) =====
class CBData:
    def __init__(self, timesteps, ep_rewards, ep_lengths):
        self.timesteps  = list(timesteps)
        self.ep_rewards = list(ep_rewards)
        self.ep_lengths = list(ep_lengths)
    def smoothed(self, window: int = 20):
        import numpy as _np
        r = _np.array(self.ep_rewards)
        if len(r) < window:
            return r
        k = _np.ones(window) / window
        return _np.convolve(r, k, mode="valid")
